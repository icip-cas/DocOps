import json
import os
import re
import sys
from pathlib import Path
from datetime import date, datetime

import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from pypdf import PdfReader

sys.path.insert(0, str(Path(__file__).parent))
from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(os.environ.get('TASK_METADATA_PATH', '/tests/task_metadata.json'))
META = json.loads(META_PATH.read_text(encoding='utf-8'))
EXPECT = META['verifier_expectations']


def _vopt_norm_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime('%Y-%m-%d %H:%M')
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    text = text.replace('\u2013', '-').replace('\u2014', '-').replace('\u2212', '-')
    text = text.replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _vopt_norm_formula(value):
    return re.sub(r'\s+', '', _vopt_norm_text(value)).upper()


def _vopt_assert_formula(actual, expected, label='formula'):
    assert _vopt_norm_formula(actual) == _vopt_norm_formula(expected), f"{label}: expected {expected!r}, found {actual!r}"


def _vopt_norm_rows(rows):
    return [[_vopt_norm_text(cell) for cell in row] for row in rows]


def _vopt_rows_equal(actual, expected):
    return _vopt_norm_rows(actual) == _vopt_norm_rows(expected)


def _vopt_rows_text(rows):
    return '\n'.join('|'.join(_vopt_norm_text(cell) for cell in row) for row in rows)


def _vopt_header_has(header, required):
    actual = {_vopt_norm_text(cell).lower() for cell in header}
    missing = [cell for cell in required if _vopt_norm_text(cell).lower() not in actual]
    assert not missing, f"Missing expected header columns {missing!r}; actual={header!r}"


def _vopt_section_text(section_part):
    return '\n'.join(p.text for p in getattr(section_part, 'paragraphs', []) if p.text is not None)


def _vopt_header_text(doc):
    return '\n'.join(_vopt_section_text(section.header) for section in doc.sections)


def _vopt_footer_text(doc):
    return '\n'.join(_vopt_section_text(section.footer) for section in doc.sections)


def _vopt_table_ref(ws, expected_name=None):
    tables = ws.tables
    if expected_name:
        for name in tables.keys():
            if str(name).lower() == str(expected_name).lower():
                return tables[name].ref
    vals = list(tables.values())
    assert vals, f"No Excel native table found on sheet {ws.title!r}"
    return vals[0].ref


def _vopt_clean_area(value):
    text = _vopt_norm_text(value).replace("'", '').replace('$', '').replace(' ', '')
    if '!' in text:
        text = text.split('!', 1)[1]
    return text.upper()


def _vopt_assert_print_area(ws, expected):
    actual = _vopt_clean_area(ws.print_area)
    target = _vopt_clean_area(expected)
    assert target in actual or actual in target, f"{ws.title}: expected print area {expected!r}, found {ws.print_area!r}"


def _vopt_freeze_pane(value):
    return getattr(value, 'coordinate', value)


def _vopt_validated_cells(ws):
    cells = set()
    for dv in ws.data_validations.dataValidation:
        for rng in dv.cells.ranges:
            text = str(rng).replace('$', '')
            try:
                min_col, min_row, max_col, max_row = range_boundaries(text)
            except ValueError:
                cells.add(text)
                continue
            if max_row > 10000:
                max_row = max(min_row, 200)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cells.add(f"{get_column_letter(col)}{row}")
    return cells


class _VoptValidationRanges(list):
    def __init__(self, ws, ranges):
        super().__init__(ranges)
        self.ws = ws

    def __contains__(self, expected):
        expected = str(expected).replace('$', '')
        if super().__contains__(expected):
            return True
        try:
            min_col, min_row, max_col, max_row = range_boundaries(expected)
        except ValueError:
            return super().__contains__(expected)
        target = {f"{get_column_letter(col)}{row}" for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1)}
        return target.issubset(_vopt_validated_cells(self.ws))


def _vopt_dv_ranges(ws):
    ranges = []
    for dv in ws.data_validations.dataValidation:
        ranges.extend(str(rng).replace('$', '') for rng in dv.cells.ranges)
    return _VoptValidationRanges(ws, ranges)


def _vopt_ordered_subset(expected, actual):
    actual_iter = iter([_vopt_norm_text(item) for item in actual])
    for item in [_vopt_norm_text(value) for value in expected]:
        for candidate in actual_iter:
            if item == candidate:
                break
        else:
            return False
    return True



def _resolve_output(kind):
    if kind == 'binder':
        return Path(os.environ.get('PDF_OUTPUT_PATH', EXPECT['binder_output']))
    if kind == 'tracker':
        return Path(os.environ.get('XLSX_OUTPUT_PATH', EXPECT['tracker_output']))
    raise KeyError(kind)


def _pdf_texts(path):
    with pdfplumber.open(str(path)) as pdf:
        return [page.extract_text() or '' for page in pdf.pages]


def _page_titles(texts):
    titles = []
    for text in texts:
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        titles.append(lines[0] if lines else '')
    return titles


def _contains_norm(text, phrase):
    return normalize_text(phrase) in normalize_text(text)


def _page_matches_title(page_text, expected_title):
    text = normalize_text(page_text)
    title = normalize_text(expected_title)
    if title in text:
        return True
    groups = {
        'Closeout Binder': ['closeout', 'binder'],
        'Open Punchlist Summary': ['punchlist'],
        'Warranty Follow-up Summary': ['warranty'],
        'Photo Evidence Appendix': ['photo', 'appendix'],
    }
    return all(token in text for token in groups.get(expected_title, [title]))


def _assert_count_phrase(full_text, labels, count):
    text = normalize_text(full_text).replace(':', ' ')
    for label in labels:
        pattern = rf'{re.escape(normalize_text(label))}\s+{count}\b'
        if re.search(pattern, text):
            return
    raise AssertionError(f"Missing count {count} for one of {labels!r}")


def _assert_pdf_required_content(full):
    text = normalize_text(full)
    _assert_count_phrase(full, ['Open Punch Items', 'Open Punchlist Items'], 2)
    _assert_count_phrase(full, ['Open Warranty Items'], 2)
    assert 'critical' in text, "binder: missing critical follow-up token 'critical'"
    assert 'pl-04' in text or 'pl 04' in text, "binder: missing critical follow-up token 'PL-04'"
    assert 'w-18' in text or 'w 18' in text, "binder: missing critical follow-up token 'W-18'"
    assert 'follow up' in text or 'follow-up' in text, "binder: missing critical follow-up wording"
    for sid in ('PL-04', 'PL-07', 'W-18', 'W-29'):
        assert normalize_text(sid) in text, f'binder: missing required source ID {sid}'


def _assert_no_forbidden_pdf_content(full):
    text = normalize_text(full)
    forbidden = [
        'internal only',
        'status draft',
        'punchlist summary draft',
        'warranty summary draft',
        'photo appendix draft',
        'draft appendix placeholder',
        'pl-11 still shown open',
        'w-22 received item carried',
    ]
    hits = [phrase for phrase in forbidden if phrase in text]
    assert not hits, f'binder: forbidden active draft/internal content remains: {hits}'


def _header_map(ws):
    return {_vopt_norm_text(cell.value).lower(): idx + 1 for idx, cell in enumerate(ws[1]) if _vopt_norm_text(cell.value)}


def _find_col(mapping, *candidates):
    for candidate in candidates:
        key = candidate.lower()
        if key in mapping:
            return mapping[key]
    for key, idx in mapping.items():
        if any(candidate.lower() in key for candidate in candidates):
            return idx
    raise AssertionError(f'Missing expected column among {candidates!r}; actual={list(mapping)!r}')


def _records_by_source_id(ws):
    mapping = _header_map(ws)
    source_col = _find_col(mapping, 'Source ID', 'Item ID', 'Warranty ID')
    records = {}
    for row in range(2, ws.max_row + 1):
        source_id = _vopt_norm_text(ws.cell(row, source_col).value)
        if source_id:
            records[source_id] = {'_row': row, '_text': _vopt_norm_text(' | '.join(_vopt_norm_text(ws.cell(row, c).value) for c in range(1, ws.max_column + 1)))}
    return records


def _assert_tracker_semantics(ws):
    records = _records_by_source_id(ws)
    expected_ids = ['PL-04', 'PL-07', 'W-18', 'W-29']
    assert set(records) == set(expected_ids), f'Unexpected follow-up IDs: {set(records)!r}'
    expected = {
        'PL-04': ['Punchlist', 'Level 2 north stair handrail finish mismatch', 'Rosa Iyer', 'Open', '2026-08-12'],
        'PL-07': ['Punchlist', 'Mechanical room label set incomplete', 'Noah Patel', 'Open', '2026-08-16'],
        'W-18': ['Warranty', 'Air handler VFD warranty letter', 'Noah Patel', '2026-08-20'],
        'W-29': ['Warranty', 'Lighting controls calibration warranty', 'Mina Zhou', '2026-08-18'],
    }
    for source_id, tokens in expected.items():
        row_text = records[source_id]['_text']
        for token in tokens:
            assert normalize_text(token) in normalize_text(row_text), f'{source_id}: missing {token!r} in row {row_text!r}'
    assert 'PL-11' not in records and 'W-22' not in records
    return records


def _row_values(ws, row, cols):
    return [ws.cell(row, c).value for c in range(1, cols + 1)]


def _direct_row_highlighted(ws, row):
    for col in range(1, ws.max_column + 1):
        rgb = cell_fill_rgb(ws.cell(row, col))
        if rgb and rgb not in ('FFFFFF', '000000', '00000000'):
            return True
    return False


def _cf_fill_present(rule):
    fill = getattr(getattr(rule, 'dxf', None), 'fill', None)
    if fill is None:
        return False
    for color in (getattr(fill, 'fgColor', None), getattr(fill, 'bgColor', None), getattr(fill, 'start_color', None)):
        if color is None:
            continue
        if getattr(color, 'rgb', None) or getattr(color, 'indexed', None) is not None or getattr(color, 'theme', None) is not None:
            return True
    return getattr(fill, 'patternType', None) is not None


def _range_covers_row(range_text, row_idx):
    refs = re.findall(r'\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?', range_text)
    for ref in refs:
        clean = ref.replace('$', '')
        if ':' not in clean:
            clean = f'{clean}:{clean}'
        try:
            _c1, r1, _c2, r2 = range_boundaries(clean)
        except ValueError:
            continue
        if r1 <= row_idx <= r2:
            return True
    return False


def _row_highlighted(ws, row):
    if _direct_row_highlighted(ws, row):
        return True
    for rng in ws.conditional_formatting:
        if not _range_covers_row(str(rng), row):
            continue
        for rule in ws.conditional_formatting[rng]:
            if _cf_fill_present(rule):
                return True
    return False


def _formula(value, label):
    assert isinstance(value, str) and value.strip().startswith('='), f'{label}: expected live formula, found {value!r}'
    return _vopt_norm_formula(value)


def _assert_binder_flag_formula(value, row):
    formula = _formula(value, f'Follow-up Tracker!I{row}')
    row_refs = {f'{col}{row}' for col in 'ABCDEFGHI'}
    assert any(ref in formula for ref in row_refs) or 'FOLLOW-UPTRACKER' in formula or 'FOLLOWUPTRACKER' in formula, (
        f'Follow-up Tracker!I{row}: formula should depend on the tracker row'
    )


def _assert_summary_formula(value, label, criteria):
    formula = _formula(value, f'Release Summary {label}')
    compact = formula.replace("'", '')
    assert 'FOLLOW-UPTRACKER' in compact or 'FOLLOWUPTRACKER' in compact, (
        f'Release Summary {label}: should reference Follow-up Tracker'
    )
    assert any(criterion.upper() in compact for criterion in criteria), (
        f'Release Summary {label}: missing expected criterion among {criteria!r}'
    )
    assert any(fn in compact for fn in ('COUNTIF', 'COUNTIFS', 'SUMPRODUCT')), (
        f'Release Summary {label}: expected a counting formula, found {value!r}'
    )


def test_outputs_exist():
    binder = _resolve_output('binder')
    tracker = _resolve_output('tracker')
    assert binder.exists(), f'Missing binder output: {binder}'
    assert tracker.exists(), f'Missing tracker output: {tracker}'
    assert binder.suffix.lower() == '.pdf'
    assert tracker.suffix.lower() == '.xlsx'


def test_pdf_binder_content_and_outline():
    binder = _resolve_output('binder')
    reader = PdfReader(str(binder))
    assert len(reader.pages) >= 4
    texts = _pdf_texts(binder)
    full = '\n'.join(texts)
    assert any(_page_matches_title(text, 'Closeout Binder') for text in texts), 'Missing closeout binder cover page'
    assert any(_page_matches_title(text, 'Open Punchlist Summary') or all(normalize_text(tok) in normalize_text(text) for tok in ('PL-04', 'PL-07')) for text in texts), 'Missing punchlist follow-up content'
    assert any(_page_matches_title(text, 'Warranty Follow-up Summary') or all(normalize_text(tok) in normalize_text(text) for tok in ('W-18', 'W-29')) for text in texts), 'Missing warranty follow-up content'
    assert any(_page_matches_title(text, 'Photo Evidence Appendix') for text in texts), 'Missing photo appendix content'
    _assert_pdf_required_content(full)
    _assert_no_forbidden_pdf_content(full)
    outline = [title for _level, title in flatten_outline(reader.outline)]
    assert len(outline) >= 3, 'PDF should have rebuilt bookmarks/outline entries'
    outline_text = ' '.join(outline)
    assert any(normalize_text(token) in normalize_text(outline_text) for token in ['Cover', 'Binder', 'Closeout']), (
        'Missing cover/binder outline category'
    )
    assert normalize_text('Photo') in normalize_text(outline_text), 'Missing outline category: Photo'
    assert (
        normalize_text('Summary') in normalize_text(outline_text)
        or (
            normalize_text('Punch') in normalize_text(outline_text)
            and normalize_text('Warranty') in normalize_text(outline_text)
        )
    ), 'Missing summary or punch/warranty outline categories'


def test_tracker_structure_and_values():
    wb = load_workbook(_resolve_output('tracker'), data_only=False)
    for sheet in EXPECT['tracker_sheet_order']:
        assert sheet in wb.sheetnames, f'Missing required sheet: {sheet}'
    ws = wb['Follow-up Tracker']
    _assert_tracker_semantics(ws)
    assert _vopt_table_ref(ws, 'closeout_followup_tracker') == EXPECT['tracker_table_ref']
    for row in range(2, 6):
        _assert_binder_flag_formula(ws[f'I{row}'].value, row)
    _assert_summary_formula(wb['Release Summary']['B3'].value, 'B3', ['Punchlist'])
    _assert_summary_formula(wb['Release Summary']['B4'].value, 'B4', ['Warranty'])
    _formula(wb['Release Summary']['B5'].value, 'Release Summary B5')
    for sheet in EXPECT['hidden_sheets']:
        assert wb[sheet].sheet_state in ('hidden', 'veryHidden'), f'{sheet} should be hidden'
    names = {dn.name for dn in wb.defined_names.values()}
    for name in EXPECT['defined_names']:
        assert name in names, f'Missing defined name: {name}'
    actual_ranges = _vopt_dv_ranges(ws)
    for expected in EXPECT['data_validation_ranges']['Follow-up Tracker']:
        assert expected in actual_ranges, f'Missing data validation range: {expected}'


def test_cross_output_consistency():
    wb = load_workbook(_resolve_output('tracker'), data_only=False)
    ws = wb['Follow-up Tracker']
    source_ids = list(_records_by_source_id(ws))
    full = '\n'.join(_pdf_texts(_resolve_output('binder')))
    for sid in source_ids:
        assert sid in full, f'{sid} missing from PDF binder'
    _assert_count_phrase(full, ['Open Punch Items', 'Open Punchlist Items'], 2)
    _assert_count_phrase(full, ['Open Warranty Items'], 2)
    assert 'PL-11' not in source_ids
    assert 'W-22' not in source_ids
