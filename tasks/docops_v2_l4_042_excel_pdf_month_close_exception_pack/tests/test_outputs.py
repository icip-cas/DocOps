import json
import os
import re
import sys
from pathlib import Path
from datetime import date, datetime

import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from pypdf import PdfReader

sys.path.insert(0, str(Path(__file__).parent))
from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(os.environ.get('TASK_METADATA_PATH', '/tests/task_metadata.json'))
META = json.loads(META_PATH.read_text(encoding='utf-8'))
INPUT_PATH = Path(os.environ.get('INPUT_PATH', META['input_path']))
EXPECT = META['verifier_expectations']


def _vopt_norm_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime('%Y-%m-%d %H:%M')
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    text = text.replace('\u2013', '-').replace('\u2014', '-').replace('\u2212', '-')
    text = text.replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _vopt_norm_formula(value):
    return re.sub(r'\s+', '', _vopt_norm_text(value)).upper()


def _vopt_assert_formula(actual, expected, label='formula'):
    assert _vopt_norm_formula(actual) == _vopt_norm_formula(expected), f"{label}: expected {expected!r}, found {actual!r}"


def _vopt_norm_rows(rows):
    return [[_vopt_norm_text(cell) for cell in row] for row in rows]


def _vopt_rows_equal(actual, expected):
    return _vopt_norm_rows(actual) == _vopt_norm_rows(expected)


def _vopt_rows_text(rows):
    return '\n'.join('|'.join(_vopt_norm_text(cell) for cell in row) for row in rows)


def _vopt_header_has(header, required):
    actual = {_vopt_norm_text(cell).lower() for cell in header}
    missing = [cell for cell in required if _vopt_norm_text(cell).lower() not in actual]
    assert not missing, f"Missing expected header columns {missing!r}; actual={header!r}"


def _vopt_section_text(section_part):
    return '\n'.join(p.text for p in getattr(section_part, 'paragraphs', []) if p.text is not None)


def _vopt_header_text(doc):
    return '\n'.join(_vopt_section_text(section.header) for section in doc.sections)


def _vopt_footer_text(doc):
    return '\n'.join(_vopt_section_text(section.footer) for section in doc.sections)


def _vopt_table_ref(ws, expected_name=None):
    tables = ws.tables
    if expected_name:
        for name in tables.keys():
            if str(name).lower() == str(expected_name).lower():
                return tables[name].ref
    vals = list(tables.values())
    assert vals, f"No Excel native table found on sheet {ws.title!r}"
    return vals[0].ref


def _vopt_clean_area(value):
    text = _vopt_norm_text(value).replace("'", '').replace('$', '').replace(' ', '')
    if '!' in text:
        text = text.split('!', 1)[1]
    return text.upper()


def _vopt_assert_print_area(ws, expected):
    actual = _vopt_clean_area(ws.print_area)
    target = _vopt_clean_area(expected)
    assert target in actual or actual in target, f"{ws.title}: expected print area {expected!r}, found {ws.print_area!r}"


def _vopt_freeze_pane(value):
    return getattr(value, 'coordinate', value)


def _vopt_validated_cells(ws):
    cells = set()
    for dv in ws.data_validations.dataValidation:
        for rng in dv.cells.ranges:
            text = str(rng).replace('$', '')
            try:
                min_col, min_row, max_col, max_row = range_boundaries(text)
            except ValueError:
                cells.add(text)
                continue
            if max_row > 10000:
                max_row = max(min_row, 200)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cells.add(f"{get_column_letter(col)}{row}")
    return cells


class _VoptValidationRanges(list):
    def __init__(self, ws, ranges):
        super().__init__(ranges)
        self.ws = ws

    def __contains__(self, expected):
        expected = str(expected).replace('$', '')
        if super().__contains__(expected):
            return True
        try:
            min_col, min_row, max_col, max_row = range_boundaries(expected)
        except ValueError:
            return super().__contains__(expected)
        target = {f"{get_column_letter(col)}{row}" for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1)}
        return target.issubset(_vopt_validated_cells(self.ws))


def _vopt_dv_ranges(ws):
    ranges = []
    for dv in ws.data_validations.dataValidation:
        ranges.extend(str(rng).replace('$', '') for rng in dv.cells.ranges)
    return _VoptValidationRanges(ws, ranges)


def _vopt_ordered_subset(expected, actual):
    actual_iter = iter([_vopt_norm_text(item) for item in actual])
    for item in [_vopt_norm_text(value) for value in expected]:
        for candidate in actual_iter:
            if item == candidate:
                break
        else:
            return False
    return True



def _resolve_output(kind):
    if kind == 'workbook':
        return Path(os.environ.get('XLSX_OUTPUT_PATH', EXPECT['workbook_output']))
    if kind == 'packet':
        return Path(os.environ.get('PDF_OUTPUT_PATH', EXPECT['packet_output']))
    raise KeyError(kind)


def _row_values(ws, row, cols):
    return [ws.cell(row, c).value for c in range(1, cols + 1)]


def _headers(ws):
    return [_vopt_norm_text(cell.value) for cell in ws[1]]


def _header_map(ws):
    return {name.lower(): idx + 1 for idx, name in enumerate(_headers(ws)) if name}


def _find_col(mapping, *candidates):
    for candidate in candidates:
        key = candidate.lower()
        if key in mapping:
            return mapping[key]
    for key, idx in mapping.items():
        if any(candidate.lower() in key for candidate in candidates):
            return idx
    raise AssertionError(f'Missing expected column among {candidates!r}; actual columns={list(mapping)!r}')


def _table_body_rows(ws):
    rows = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if any(value is not None and _vopt_norm_text(value) for value in values):
            rows.append((row, values))
    return rows


def _parse_money(value):
    if isinstance(value, (int, float)):
        return float(value)
    text = _vopt_norm_text(value)
    match = re.search(r'\(?\$?\s*([0-9][0-9,]*(?:\.\d+)?)\)?', text)
    assert match, f'Expected money amount, found {value!r}'
    return float(match.group(1).replace(',', ''))


def _formula_has(value, label, *tokens):
    norm = _vopt_norm_formula(value)
    assert norm.startswith('='), f'{label}: expected live formula, found {value!r}'
    for token in tokens:
        assert _vopt_norm_formula(token) in norm, f'{label}: expected formula to reference {token!r}, found {value!r}'


def _clean_transaction_records(ws):
    mapping = _header_map(ws)
    review_col = None
    try:
        review_col = _find_col(mapping, 'Review Flag', 'Review', 'Review Status')
    except AssertionError:
        review_col = None
    cols = {
        'txn': _find_col(mapping, 'Txn ID', 'Transaction ID'),
        'date': _find_col(mapping, 'Date'),
        'vendor': _find_col(mapping, 'Vendor'),
        'category': _find_col(mapping, 'Final Category', 'Category'),
        'receipt': _find_col(mapping, 'Receipt ID'),
        'amount': _find_col(mapping, 'Amount'),
        'status': _find_col(mapping, 'Receipt Status', 'Status'),
    }
    if review_col:
        cols['review'] = review_col
    records = {}
    for row_idx, _values in _table_body_rows(ws):
        txn = _vopt_norm_text(ws.cell(row_idx, cols['txn']).value)
        if not txn:
            continue
        records[txn] = {
            key: ws.cell(row_idx, col).value for key, col in cols.items()
        }
        records[txn]['row_idx'] = row_idx
    return records


def _exception_text_rows(ws):
    rows = []
    for row_idx, values in _table_body_rows(ws):
        text = ' '.join(_vopt_norm_text(value) for value in values if _vopt_norm_text(value))
        if not text:
            continue
        if 'total' in normalize_text(text) or 'count' in normalize_text(text):
            continue
        rows.append((row_idx, text, values))
    return rows


def _assert_exception_row(rows, label, required_tokens, amount):
    matches = []
    for _row_idx, text, values in rows:
        norm = normalize_text(text)
        if all(normalize_text(token) in norm for token in required_tokens):
            row_amounts = []
            for value in values:
                try:
                    row_amounts.append(_parse_money(value))
                except AssertionError:
                    continue
            if any(abs(candidate - amount) < 0.01 for candidate in row_amounts):
                matches.append(text)
    assert matches, f'{label}: expected tokens {required_tokens!r} and amount ${amount:,.0f}; rows={[row[1] for row in rows]!r}'


def _extract_workbook_exception_amounts(ws):
    rows = _exception_text_rows(ws)
    expected = []
    for _label, tokens, amount in _expected_exception_specs():
        _assert_exception_row(rows, _label, tokens, amount)
        expected.append(amount)
    return expected


def _expected_exception_specs():
    return [
        ('duplicate exception', ['RC-1002'], 1180),
        ('missing receipt exception', ['RC-1004'], 875),
        ('policy exception', ['TRX-107'], 1975),
    ]


def _assert_review_flag(value, should_review, txn):
    norm = normalize_text(_vopt_norm_text(value))
    if should_review:
        assert norm in {'review', 'yes', 'y', 'true', 'needs review'}, f'{txn}: expected review flag, found {value!r}'
    else:
        assert norm in {'ok', 'no', 'n', 'false', 'cleared', 'clear', ''}, f'{txn}: expected non-review flag, found {value!r}'


def _pdf_texts(path):
    with pdfplumber.open(str(path)) as pdf:
        return [page.extract_text() or '' for page in pdf.pages]


def _page_titles(texts):
    out = []
    for text in texts:
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        out.append(lines[0] if lines else '')
    return out


def _row_highlighted(ws, row):
    for col in range(1, ws.max_column + 1):
        rgb = cell_fill_rgb(ws.cell(row, col))
        if rgb and rgb not in ('FFFFFF', '000000', '00000000'):
            return True
    return False


def _has_exception_count(text):
    norm = normalize_text(text).replace(':', ' ')
    patterns = [
        r'\bexception count\s*3\b',
        r'\bopen exception count\s*3\b',
        r'\bopen exceptions\s*3\b',
        r'\btotal open exceptions\s*3\b',
    ]
    return any(re.search(pattern, norm) for pattern in patterns)


def _outline_text(reader):
    return ' '.join(title for _level, title in flatten_outline(reader.outline))


def _assert_pdf_outline(reader):
    titles = [title for _level, title in flatten_outline(reader.outline)]
    text = normalize_text(' '.join(titles))
    assert 'cover' in text or 'packet' in text, f'Missing cover/packet bookmark: {titles!r}'
    exception_titles = [
        title for title in titles
        if any(token in normalize_text(title) for token in ('ex-', 'exc-', 'exception', 'duplicate', 'missing', 'policy'))
    ]
    assert len(exception_titles) >= 3, f'Missing bookmarks for the three exception pages: {titles!r}'


def test_outputs_exist():
    workbook = _resolve_output('workbook')
    packet = _resolve_output('packet')
    assert workbook.exists(), f'Missing workbook output: {workbook}'
    assert packet.exists(), f'Missing PDF packet output: {packet}'
    assert workbook.suffix.lower() == '.xlsx'
    assert packet.suffix.lower() == '.pdf'


def test_workbook_structure_formulas_and_rows():
    wb = load_workbook(_resolve_output('workbook'), data_only=False)
    for sheet in EXPECT['sheet_order']:
        assert sheet in wb.sheetnames, f"Missing expected sheet {sheet!r}; found {wb.sheetnames!r}"
    summary = wb['Close Summary']
    _formula_has(summary['B3'].value, 'B3 revenue formula', 'Clean Transactions', 'Revenue')
    _formula_has(summary['B4'].value, 'B4 expense formula', 'Clean Transactions')
    _formula_has(summary['B5'].value, 'B5 net formula', 'B3', 'B4')
    _formula_has(summary['B7'].value, 'B7 exception count formula', 'Exceptions')
    _formula_has(summary['B8'].value, 'B8 exception amount formula', 'Exceptions')

    clean_ws = wb['Clean Transactions']
    records = _clean_transaction_records(clean_ws)
    assert set(records) == {'TRX-101', 'TRX-102', 'TRX-103', 'TRX-104', 'TRX-105', 'TRX-106', 'TRX-107'}, (
        f'Unexpected Clean Transactions IDs: {set(records)!r}'
    )
    expected_clean = {
        'TRX-101': ('Revenue', 'RC-1001', 12400, 'Present', False),
        'TRX-102': ('Expense', 'RC-1002', -1180, 'Duplicate', True),
        'TRX-103': ('Expense', 'RC-1002', -1180, 'Duplicate', True),
        'TRX-104': ('Expense', 'RC-1004', -875, 'Missing', True),
        'TRX-105': ('Revenue', 'RC-1005', 7400, 'Present', False),
        'TRX-106': ('Revenue', 'RC-1006', 6100, 'Present', False),
        'TRX-107': ('Consulting', 'RC-1007', -1975, 'Policy Exception', True),
    }
    for txn, (category, receipt, amount, status, should_review) in expected_clean.items():
        actual = records[txn]
        assert _vopt_norm_text(actual['category']) == category, f'{txn}: wrong category {actual["category"]!r}'
        assert _vopt_norm_text(actual['receipt']) == receipt, f'{txn}: wrong receipt {actual["receipt"]!r}'
        assert abs(float(actual['amount']) - amount) < 0.01, f'{txn}: wrong amount {actual["amount"]!r}'
        assert _vopt_norm_text(actual['status']) == status, f'{txn}: wrong receipt status {actual["status"]!r}'
        if 'review' in actual:
            _assert_review_flag(actual['review'], should_review, txn)
        elif should_review:
            assert _row_highlighted(clean_ws, actual['row_idx']), f'{txn}: expected highlighted review row when review flag column is absent'
    exc_ws = wb['Exceptions']
    assert len(_exception_text_rows(exc_ws)) == 3, f'Expected exactly three exception rows, found {_exception_text_rows(exc_ws)!r}'
    _extract_workbook_exception_amounts(exc_ws)

    assert _vopt_table_ref(clean_ws, 'CleanTransactions').upper().endswith('8'), 'CleanTransactions must include all seven rows'
    assert _vopt_table_ref(exc_ws, 'CloseExceptions').upper().endswith('4'), 'CloseExceptions must include exactly three exception rows'
    for sheet in EXPECT['hidden_sheets']:
        assert wb[sheet].sheet_state in ('hidden', 'veryHidden'), f'{sheet} should be hidden'
    names = {dn.name for dn in wb.defined_names.values()}
    for name in EXPECT['defined_names']:
        assert name in names, f'Missing defined name: {name}'
    actual_ranges = _vopt_dv_ranges(clean_ws)
    for expected in EXPECT['data_validation_ranges']['Clean Transactions']:
        assert expected in actual_ranges, f'Missing data validation range: {expected}'
    for spec in EXPECT['highlight_rows']:
        assert _row_highlighted(wb[spec['sheet']], spec['row']), f"Expected highlighted row {spec['sheet']}!{spec['row']}"


def test_pdf_packet_content_and_outline():
    packet = _resolve_output('packet')
    reader = PdfReader(str(packet))
    assert len(reader.pages) >= 4, f'Expected at least cover plus three exception pages, found {len(reader.pages)}'
    texts = _pdf_texts(packet)
    full_text = '\n'.join(texts)
    require_all(full_text, ['Exception Packet', 'Total Exception Amount'], 'PDF packet')
    assert _has_exception_count(full_text), 'PDF packet: missing exception count 3'
    for label, tokens, amount in _expected_exception_specs():
        require_all(full_text, tokens, f'PDF {label}')
        assert f"${amount:,}" in full_text or f"${amount:,.2f}" in full_text, f'PDF {label}: missing amount ${amount:,.0f}'
    forbid_any(full_text, EXPECT['pdf_forbidden'], 'PDF packet')
    _assert_pdf_outline(reader)


def test_cross_output_exception_consistency():
    wb = load_workbook(_resolve_output('workbook'), data_only=False)
    exc_ws = wb['Exceptions']
    amounts = _extract_workbook_exception_amounts(exc_ws)
    total = int(sum(amounts))
    texts = '\n'.join(_pdf_texts(_resolve_output('packet')))
    for _label, tokens, amount in _expected_exception_specs():
        require_all(texts, tokens, f'PDF {_label}')
        assert f"${amount:,}" in texts or f"${amount:,.2f}" in texts, f'PDF {_label}: missing amount ${amount:,.0f}'
    assert f"${total:,}" in texts, f'Workbook exception total ${total:,} missing from PDF packet'
    assert _has_exception_count(texts)
