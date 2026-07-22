import json
import os
import re
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from pypdf import PdfReader

from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]


def _norm(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value).replace("\u2013", "-").replace("\u2014", "-").replace("\xa0", " ")).strip()


def _norm_formula(value):
    return re.sub(r"\s+", "", _norm(value)).upper()


def _norm_key(value):
    return re.sub(r"[^a-z0-9]+", "", _norm(value).casefold())


def _title_key(value):
    return re.sub(r"[^a-z0-9]+", "", re.sub(r"^\s*the\s+", "", _norm(value), flags=re.I).casefold())


def _time_minutes(value):
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    text = _norm(value).upper()
    for fmt in ("%I:%M %p", "%H:%M", "%I:%M%p"):
        try:
            t = datetime.strptime(text, fmt).time()
            return t.hour * 60 + t.minute
        except ValueError:
            pass
    raise AssertionError(f"Could not parse time value: {value!r}")


def _time_label(minutes):
    hour = (minutes // 60) % 24
    minute = minutes % 60
    suffix = "AM" if hour < 12 else "PM"
    display_hour = hour % 12 or 12
    return f"{display_hour}:{minute:02d} {suffix}"


def _path(kind):
    env = {"pdf": "PDF_OUTPUT_PATH", "xlsx": "XLSX_OUTPUT_PATH"}[kind]
    key = {"pdf": "pdf_output", "xlsx": "xlsx_output"}[kind]
    return Path(os.environ.get(env, EXPECT[key]))


def _pdf_text(path):
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages), len(reader.pages)


def _expected_header(rows):
    return [_norm(c) for c in rows[0]]


def _sheet(wb, expected_name):
    _assert_required_sheets(wb, [expected_name])
    return wb[expected_name]


def _table(ws, expected_name):
    for name in ws.tables.keys():
        if str(name).lower() == str(expected_name).lower():
            return ws.tables[name]
    raise AssertionError(f"Missing table {expected_name!r} on {ws.title!r}")


def _table_bounds(table):
    return range_boundaries(table.ref.replace("$", ""))


def _table_headers(ws, table):
    min_col, min_row, max_col, _ = _table_bounds(table)
    return [_norm(ws.cell(min_row, col).value) for col in range(min_col, max_col + 1)]


def _table_header_map(ws, table):
    min_col, _, _, _ = _table_bounds(table)
    return {_norm_key(header): idx for idx, header in enumerate(_table_headers(ws, table), start=min_col)}


def _require_headers(ws, table, expected_headers):
    actual = {_norm_key(header) for header in _table_headers(ws, table)}
    missing = [header for header in expected_headers if _norm_key(header) not in actual]
    assert not missing, f"{ws.title}: table {table.name} missing headers {missing}; found {_table_headers(ws, table)}"


def _table_has_size(table, min_data_rows, min_cols):
    min_col, min_row, max_col, max_row = _table_bounds(table)
    assert max_col - min_col + 1 >= min_cols, f"{table.name}: expected at least {min_cols} columns"
    assert max_row - min_row >= min_data_rows, f"{table.name}: expected at least {min_data_rows} data rows"


def _table_rows(ws, table):
    _, min_row, _, max_row = _table_bounds(table)
    header_map = _table_header_map(ws, table)
    rows = []
    for row_idx in range(min_row + 1, max_row + 1):
        row = {"__row_idx": row_idx}
        for key, col_idx in header_map.items():
            row[key] = ws.cell(row_idx, col_idx)
        if any(cell.value is not None for key, cell in row.items() if key != "__row_idx"):
            rows.append(row)
    return rows


def _cell_by_header(row, header):
    key = _norm_key(header)
    assert key in row, f"Missing header {header!r}"
    return row[key]


def _value_matches(actual, expected):
    actual_text = _norm(actual)
    expected_text = _norm(expected)
    if actual_text.casefold() == expected_text.casefold():
        return True
    if len(expected_text) >= 12 and expected_text.casefold() in actual_text.casefold():
        return True
    return False


FIELD_ALIASES = {
    "festival": {"festival", "festival name"},
    "festival date": {"festival date", "date"},
    "public screenings": {"public screenings", "total screenings", "screening count", "screenings"},
    "constraint checks passed": {"constraint checks passed", "constraint pass count", "constraints passed"},
    "access notes included": {"access notes included", "access-note count", "access notes", "included access notes"},
    "private notes excluded": {"private notes excluded", "private notes excluded count", "private exclusions"},
    "program ready": {"program ready", "programready", "ready status"},
}


def _key_value_matches(actual, expected):
    if _value_matches(actual, expected):
        return True
    expected_text = _norm(expected).casefold()
    aliases = {_norm_key(item) for item in FIELD_ALIASES.get(expected_text, set())}
    return _norm_key(actual) in aliases and _norm_key(expected) in aliases


def _find_row(rows, key_header, key_value):
    for row in rows:
        if _key_value_matches(_cell_by_header(row, key_header).value, key_value):
            return row
    raise AssertionError(f"Could not find row with {key_header!r}={key_value!r}")


def _assert_row_contains(rows, headers, expected_row, key_header=None, skip_headers=()):
    key_header = key_header or headers[0]
    row = _find_row(rows, key_header, expected_row[headers.index(key_header)])
    for header, expected_value in zip(headers, expected_row):
        if header in skip_headers:
            continue
        actual = _cell_by_header(row, header).value
        assert _value_matches(actual, expected_value), f"{header}: expected {expected_value!r}, found {actual!r}"
    return row


def _range_covered(ws, expected):
    min_col, min_row, max_col, max_row = range_boundaries(expected.replace("$", ""))
    target = {f"{get_column_letter(col)}{row}" for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1)}
    cells = set()
    for dv in ws.data_validations.dataValidation:
        for rng in dv.cells.ranges:
            min_c, min_r, max_c, max_r = range_boundaries(str(rng).replace("$", ""))
            cells.update(f"{get_column_letter(col)}{row}" for row in range(min_r, max_r + 1) for col in range(min_c, max_c + 1))
    return target.issubset(cells)


def _clean_area(value):
    if isinstance(value, (list, tuple)):
        value = ",".join(str(item) for item in value)
    text = _norm(value).replace("'", "").replace("$", "").replace(" ", "")
    if "!" in text:
        text = text.split("!", 1)[1]
    return text.upper()


def _sheet_text(wb, sheets):
    _assert_required_sheets(wb, sheets)
    parts = []
    for sheet in sheets:
        for row in wb[sheet].iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
    return "\n".join(parts)


def _assert_required_sheets(wb, sheets):
    missing = [sheet for sheet in sheets if sheet not in wb.sheetnames]
    assert not missing, f"Missing required workbook sheets: {missing}; found {wb.sheetnames}"


def _formula_refs(value, *refs):
    formula = _norm_formula(value)
    assert formula.startswith("="), f"expected formula, got {value!r}"
    for ref in refs:
        assert _norm_formula(ref) in formula, f"formula {value!r} does not reference {ref}"


def _formula_targets(value, sheet_name, table_name, status_or_count=None):
    formula = _norm_formula(value)
    assert formula.startswith("="), f"expected formula, got {value!r}"
    refs_target = sheet_name.replace(" ", "").upper() in formula.replace("'", "") or table_name.upper() in formula
    assert refs_target, f"formula {value!r} must reference {sheet_name} or {table_name}"
    if status_or_count:
        assert status_or_count.upper() in formula, f"formula {value!r} must count/check {status_or_count}"


def _block_formula_ok(value, row, header_map):
    formula = _norm_formula(value)
    assert formula.startswith("="), f"block minutes must be formula-backed, got {value!r}"
    refs = []
    for header in ("Runtime Min", "Q&A Min", "Turnover Min"):
        col = header_map[_norm_key(header)]
        refs.append(f"{get_column_letter(col)}{row['__row_idx']}".upper())
    structured_tokens = ("RUNTIMEMIN", "QAMIN", "TURNOVERMIN")
    assert all(ref in formula for ref in refs) or all(token in formula for token in structured_tokens), (
        f"block formula {value!r} must reference runtime, Q&A, and turnover"
    )


def _defined_name_names(wb):
    try:
        values = wb.defined_names.values()
    except AttributeError:
        values = getattr(wb.defined_names, "definedName", [])
    return {dn.name for dn in values if getattr(dn, "name", None)}


def _ready_formula_ok(value):
    formula = _norm_formula(value)
    assert formula.startswith("="), "program-ready control must be a formula"
    for token in ("READY", "HOLD"):
        assert token in formula, f"program-ready formula missing {token}"
    assert "IF" in formula or "IFS" in formula, "program-ready formula must be conditional"
    assert "AND" in formula or formula.count("=") >= 4, "program-ready formula must combine all readiness checks"
    assert any(token in formula for token in ("B4=6", "SCREENINGCOUNT=6", "PUBLICSCREENINGS", "SCREENINGS")), "program-ready formula must check six screenings"
    assert any(token in formula for token in ("B5=6", "CONSTRAINTPASSCOUNT=6", "CONSTRAINT")), "program-ready formula must check six passed constraints"
    assert any(token in formula for token in ("B6=5", "ACCESS", "INCLUDED")), "program-ready formula must check five included access notes"
    assert any(token in formula for token in ("B7=6", "PRIVATE", "EXCLUDED")), "program-ready formula must check six excluded private notes"


def _status_validation_ok(ws, table_name, status_header):
    table = _table(ws, table_name)
    _, min_row, _, max_row = _table_bounds(table)
    col_idx = _table_header_map(ws, table)[_norm_key(status_header)]
    target = f"{get_column_letter(col_idx)}{min_row + 1}:{get_column_letter(col_idx)}{max_row}"
    assert _range_covered(ws, target), f"{ws.title}: missing validation over {target}"


def _summary_rows(wb):
    ws = _sheet(wb, "Program Summary")
    table = _table(ws, "tblProgramSummary")
    return _table_rows(ws, table)


def _summary_value_cell(wb, field):
    return _cell_by_header(_find_row(_summary_rows(wb), "Field", field), "Value")


def _defined_name_cell(wb, name):
    defined = wb.defined_names.get(name)
    if defined is None:
        return None
    for sheet_name, coord in defined.destinations:
        if sheet_name in wb.sheetnames:
            target = wb[sheet_name][coord]
            if isinstance(target, tuple):
                return target[0][0]
            return target
    return None


def _summary_or_named_formula(wb, field, defined_name=None):
    try:
        return _summary_value_cell(wb, field).value
    except AssertionError:
        if defined_name:
            cell = _defined_name_cell(wb, defined_name)
            if cell is not None:
                return cell.value
        raise


def _row_text(row):
    return " ".join(_norm(cell.value) for key, cell in row.items() if key != "__row_idx" and cell.value is not None)


def _status_ok(value, expected="Pass"):
    actual = _norm(value).casefold()
    return actual in {_norm(expected).casefold(), "ready", "approved", "ok", "complete", "completed", "yes"}


def _table_key(sheet):
    return {
        "Program Summary": "summary_rows",
        "Screening Schedule": "schedule_rows",
        "Constraint QA": "constraint_rows",
        "Accessibility Map": "access_rows",
        "Public Copy": "public_copy_rows",
        "Private Notes": "private_rows",
    }[sheet]


def _pdf_screening_times(text):
    titles = [row[4] for row in EXPECT["schedule_rows"][1:]]
    entries = []
    for title in titles:
        match = re.search(re.escape(title), text, flags=re.I)
        assert match, f"PDF missing screening title {title!r}"
        window = text[max(0, match.start() - 90): min(len(text), match.end() + 90)]
        times = re.findall(r"\b(?:1[0-2]|0?[1-9]):[0-5][0-9]\s*(?:AM|PM)\b", window, flags=re.I)
        assert times, f"PDF screening {title!r} has no nearby public start time"
        title_pos = match.start() - max(0, match.start() - 90)
        best = min(times, key=lambda t: abs(window.upper().find(t.upper()) - title_pos))
        entries.append((match.start(), _time_minutes(best), title, best))
    entries.sort()
    return entries


def _schedule_records(ws):
    table = _table(ws, "tblScreeningSchedule")
    _require_headers(ws, table, _expected_header(EXPECT["schedule_rows"]))
    header_map = _table_header_map(ws, table)
    records = []
    for row in _table_rows(ws, table):
        title = _norm(_cell_by_header(row, "Title").value)
        if not title:
            continue
        start = _time_minutes(_cell_by_header(row, "Start").value)
        end_value = _cell_by_header(row, "End").value
        computed_end = start + int(_cell_by_header(row, "Runtime Min").value) + int(_cell_by_header(row, "Q&A Min").value or 0) + int(_cell_by_header(row, "Turnover Min").value or 0)
        if isinstance(end_value, str) and end_value.strip().startswith("="):
            end_minutes = computed_end
        else:
            end_minutes = _time_minutes(end_value)
        records.append({
            "row": row,
            "slot": _cell_by_header(row, "Slot").value,
            "date": _norm(_cell_by_header(row, "Date").value),
            "venue": _norm(_cell_by_header(row, "Venue").value),
            "start": start,
            "title": title,
            "runtime": int(_cell_by_header(row, "Runtime Min").value),
            "qa": int(_cell_by_header(row, "Q&A Min").value or 0),
            "turnover": int(_cell_by_header(row, "Turnover Min").value or 0),
            "block_formula": _cell_by_header(row, "Block Min").value,
            "end": end_minutes,
            "note": _norm(_cell_by_header(row, "Public Access Note").value),
            "header_map": header_map,
        })
    return records


def _validate_schedule_records(records):
    expected_runtime = {_title_key(row[4]): row[5] for row in EXPECT["schedule_rows"][1:]}
    assert {_title_key(r["title"]) for r in records} == set(expected_runtime)
    starts = [r["start"] for r in records]
    assert starts == sorted(starts), "workbook screening rows must be chronological"
    for r in records:
        assert r["date"] == "2026-10-03"
        assert r["venue"] in {"Harbor One", "Studio B"}
        assert r["runtime"] == expected_runtime[_title_key(r["title"])]
        _block_formula_ok(r["block_formula"], r["row"], r["header_map"])
        assert r["end"] == r["start"] + r["runtime"] + r["qa"] + r["turnover"]
        title_key = _title_key(r["title"])
        if title_key in {"glassorchard", "nightferry"}:
            assert r["qa"] > 0
        if title_key == "glassorchard":
            assert r["start"] >= 13 * 60, "Glass Orchard Q&A must be after Mira Chen arrives at 1:00 PM"
            assert normalize_text("Mira Chen") in normalize_text(r["note"]) or r["qa"] > 0
        if title_key == "nightferry":
            assert normalize_text("Live captioning for Q&A") in normalize_text(r["note"])
        if title_key == "paperkites":
            assert normalize_text("Audio description") in normalize_text(r["note"])
        if title_key == "saltroom":
            assert normalize_text("low-frequency sound") in normalize_text(r["note"])
    for venue in {"Harbor One", "Studio B"}:
        venue_records = sorted([r for r in records if r["venue"] == venue], key=lambda r: r["start"])
        for previous, current in zip(venue_records, venue_records[1:]):
            assert current["start"] >= previous["end"], (
                f"{venue}: {current['title']} starts at {_time_label(current['start'])} before "
                f"{previous['title']} block ends at {_time_label(previous['end'])}"
            )


def test_outputs_exist():
    assert _path("pdf").exists()
    assert _path("xlsx").exists()
    assert _path("pdf").suffix.lower() == ".pdf"
    assert _path("xlsx").suffix.lower() == ".xlsx"


def test_pdf_public_program_content_privacy():
    text, pages = _pdf_text(_path("pdf"))
    assert 1 <= pages <= 3
    require_all(text, EXPECT["pdf_required"], "public program pdf")
    forbid_any(text, EXPECT["pdf_forbidden"], "public program pdf")
    entries = _pdf_screening_times(text)
    times = [minutes for _, minutes, _, _ in entries]
    debug_entries = [(title, label) for _, _, title, label in entries]
    assert times == sorted(times), f"screenings are not in chronological order: {debug_entries}"


def test_workbook_structure_formulas_controls():
    wb = load_workbook(_path("xlsx"), data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    for sheet, (table_name, _ref) in EXPECT["tables"].items():
        ws = _sheet(wb, sheet)
        table = _table(ws, table_name)
        expected_rows = EXPECT[_table_key(sheet)]
        _require_headers(ws, table, _expected_header(expected_rows))
        _table_has_size(table, len(expected_rows) - 1, len(expected_rows[0]))
    for sheet in EXPECT["hidden_sheets"]:
        assert _sheet(wb, sheet).sheet_state in ("hidden", "veryHidden")
    names = _defined_name_names(wb)
    for name in EXPECT["defined_names"]:
        assert name in names
    _formula_targets(_summary_or_named_formula(wb, "Public screenings", "ScreeningCount"), "Screening Schedule", "tblScreeningSchedule")
    ready_formula = _summary_or_named_formula(wb, "Program ready", "ProgramReady")
    try:
        _formula_targets(_summary_value_cell(wb, "Constraint checks passed").value, "Constraint QA", "tblConstraintQA", "Pass")
    except AssertionError:
        _formula_targets(ready_formula, "Constraint QA", "tblConstraintQA", "Pass")
    try:
        _formula_targets(_summary_value_cell(wb, "Access notes included").value, "Accessibility Map", "tblAccessibilityMap", "Yes")
    except AssertionError:
        _formula_targets(ready_formula, "Accessibility Map", "tblAccessibilityMap", "Yes")
    _ready_formula_ok(ready_formula)
    for record in _schedule_records(_sheet(wb, "Screening Schedule")):
        _block_formula_ok(record["block_formula"], record["row"], record["header_map"])
    _status_validation_ok(_sheet(wb, "Constraint QA"), "tblConstraintQA", "Status")
    _status_validation_ok(_sheet(wb, "Accessibility Map"), "tblAccessibilityMap", "Include")
    _status_validation_ok(_sheet(wb, "Public Copy"), "tblPublicCopy", "Status")
    for sheet, area in EXPECT["print_areas"].items():
        actual = _clean_area(_sheet(wb, sheet).print_area)
        target = _clean_area(area)
        if sheet == "Constraint QA" and actual == "A1:D8":
            continue
        assert target in actual or actual in target


def test_workbook_values_privacy():
    wb = load_workbook(_path("xlsx"), data_only=False)
    _assert_required_sheets(wb, EXPECT["sheet_order"])
    summary_rows = _table_rows(_sheet(wb, "Program Summary"), _table(_sheet(wb, "Program Summary"), "tblProgramSummary"))
    summary_headers = EXPECT["summary_rows"][0]
    for expected_row in EXPECT["summary_rows"][1:3]:
        row = _find_row(summary_rows, "Field", expected_row[0])
        if expected_row[0] == "Festival":
            assert _value_matches(_cell_by_header(row, "Value").value, expected_row[1])
        if expected_row[0] == "Festival date":
            date_text = _norm(_cell_by_header(row, "Value").value).casefold()
            assert "2026" in date_text and ("10" in date_text or "october" in date_text)
    _validate_schedule_records(_schedule_records(_sheet(wb, "Screening Schedule")))
    constraint_rows = _table_rows(_sheet(wb, "Constraint QA"), _table(_sheet(wb, "Constraint QA"), "tblConstraintQA"))
    constraint_text = "\n".join(_row_text(row).casefold() for row in constraint_rows)
    for tokens in (
        ("harbor one", "turnaround"),
        ("studio b", "conflict"),
        ("mira chen", "availability"),
        ("night ferry", "caption"),
        ("paper kites", "audio description"),
    ):
        assert all(token in constraint_text for token in tokens), f"Constraint QA missing {' '.join(tokens)}"
    assert sum(1 for row in constraint_rows if _status_ok(_cell_by_header(row, "Status").value, "Pass")) >= 6
    access_text = _sheet_text(wb, ["Accessibility Map"])
    for film, phrase in [
        ("Tide Tables", "Open captions"),
        ("Paper Kites", "Audio description"),
        ("Night Ferry", "Live captioning"),
        ("Salt Room", "low-frequency sound"),
        ("Closing Shorts Block", "Open captions"),
    ]:
        assert normalize_text(film) in normalize_text(access_text)
        assert normalize_text(phrase) in normalize_text(access_text)
    public_copy = _sheet_text(wb, ["Public Copy"])
    for phrase in ["Harbor Lights Micro-Festival", "Rush line opens 20 minutes before each screening", "Harbor One", "Studio B"]:
        assert normalize_text(phrase) in normalize_text(public_copy)
    public_copy_rows = _table_rows(_sheet(wb, "Public Copy"), _table(_sheet(wb, "Public Copy"), "tblPublicCopy"))
    public_copy_text = _sheet_text(wb, ["Public Copy"])
    for tokens in (("festival",), ("rush", "20"), ("harbor one", "studio b"), ("schedule", "change")):
        assert all(token in normalize_text(public_copy_text).casefold() for token in tokens)
    for row in public_copy_rows:
        assert _status_ok(_cell_by_header(row, "Status").value, "Ready")
    private_text = _sheet_text(wb, ["Private Notes"])
    for phrase in EXPECT["pdf_forbidden"]:
        assert normalize_text(phrase) in normalize_text(private_text)
    try:
        private_count = _summary_value_cell(wb, "Private notes excluded").value
        if isinstance(private_count, str) and private_count.startswith("="):
            formula = _norm_formula(private_count)
            assert "PRIVATE" in formula and any(token in formula for token in ("COUNT", "COUNTA", "ROWS", "SUBTOTAL")), "private-notes-excluded formula must count private excluded notes"
        else:
            assert _norm(private_count) == "6", "private-notes-excluded count must be 6"
    except AssertionError:
        ready_formula = _norm_formula(_summary_or_named_formula(wb, "Program ready", "ProgramReady"))
        assert "PRIVATE" in ready_formula and any(token in ready_formula for token in ("COUNT", "COUNTA", "ROWS", "SUBTOTAL"))
        private_rows = _table_rows(_sheet(wb, "Private Notes"), _table(_sheet(wb, "Private Notes"), "tblPrivateNotes"))
        assert len(private_rows) == 6, "Private Notes must retain exactly six excluded notes"
    public = _sheet_text(wb, ["Program Summary", "Screening Schedule", "Constraint QA", "Accessibility Map", "Public Copy"])
    forbid_any(public, EXPECT["pdf_forbidden"], "public workbook sheets")


def test_cross_output_consistency():
    pdf_text, _ = _pdf_text(_path("pdf"))
    wb = load_workbook(_path("xlsx"), data_only=False)
    _assert_required_sheets(wb, ["Screening Schedule", "Accessibility Map", "Public Copy", "Constraint QA", "Program Summary"])
    public = _sheet_text(wb, ["Screening Schedule", "Accessibility Map", "Public Copy"])
    for anchor in ["Tide Tables", "Glass Orchard", "Paper Kites", "Night Ferry", "Salt Room", "Closing Shorts Block", "Mira Chen"]:
        assert normalize_text(anchor) in normalize_text(pdf_text)
        assert normalize_text(anchor) in normalize_text(public)
    constraint_rows = _table_rows(_sheet(wb, "Constraint QA"), _table(_sheet(wb, "Constraint QA"), "tblConstraintQA"))
    statuses = [_norm(_cell_by_header(row, "Status").value) for row in constraint_rows]
    assert statuses.count("Pass") >= 6
    _ready_formula_ok(_summary_or_named_formula(wb, "Program ready", "ProgramReady"))
