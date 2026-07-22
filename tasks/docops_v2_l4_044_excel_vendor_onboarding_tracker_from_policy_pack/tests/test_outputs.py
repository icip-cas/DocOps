import json
import os
import re
import sys
from pathlib import Path
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

sys.path.insert(0, str(Path(__file__).parent))
from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(os.environ.get('TASK_METADATA_PATH', '/tests/task_metadata.json'))
META = json.loads(META_PATH.read_text(encoding='utf-8'))
INPUT_PATH = Path(os.environ.get('INPUT_PATH', META['input_path']))
OUTPUT_PATH = Path(os.environ.get('OUTPUT_PATH', META['output_path']))
EXPECT = META['verifier_expectations']


def _vopt_norm_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime('%Y-%m-%d %H:%M')
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    text = text.replace('\u2013', '-').replace('\u2014', '-').replace('\u2212', '-')
    text = text.replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _vopt_norm_formula(value):
    return re.sub(r'\s+', '', _vopt_norm_text(value)).upper()


def _vopt_assert_formula(actual, expected, label='formula'):
    assert _vopt_norm_formula(actual) == _vopt_norm_formula(expected), f"{label}: expected {expected!r}, found {actual!r}"


def _vopt_norm_rows(rows):
    return [[_vopt_norm_text(cell) for cell in row] for row in rows]


def _vopt_rows_equal(actual, expected):
    return _vopt_norm_rows(actual) == _vopt_norm_rows(expected)


def _vopt_rows_text(rows):
    return '\n'.join('|'.join(_vopt_norm_text(cell) for cell in row) for row in rows)


def _vopt_header_has(header, required):
    actual = {_vopt_norm_text(cell).lower() for cell in header}
    missing = [cell for cell in required if _vopt_norm_text(cell).lower() not in actual]
    assert not missing, f"Missing expected header columns {missing!r}; actual={header!r}"


def _vopt_section_text(section_part):
    return '\n'.join(p.text for p in getattr(section_part, 'paragraphs', []) if p.text is not None)


def _vopt_header_text(doc):
    return '\n'.join(_vopt_section_text(section.header) for section in doc.sections)


def _vopt_footer_text(doc):
    return '\n'.join(_vopt_section_text(section.footer) for section in doc.sections)


def _vopt_table_ref(ws, expected_name=None):
    tables = ws.tables
    if expected_name:
        for name in tables.keys():
            if str(name).lower() == str(expected_name).lower():
                return tables[name].ref
    vals = list(tables.values())
    assert vals, f"No Excel native table found on sheet {ws.title!r}"
    return vals[0].ref


def _vopt_clean_area(value):
    text = _vopt_norm_text(value).replace("'", '').replace('$', '').replace(' ', '')
    if '!' in text:
        text = text.split('!', 1)[1]
    return text.upper()


def _vopt_assert_print_area(ws, expected):
    actual = _vopt_clean_area(ws.print_area)
    target = _vopt_clean_area(expected)
    assert target in actual or actual in target, f"{ws.title}: expected print area {expected!r}, found {ws.print_area!r}"


def _vopt_freeze_pane(value):
    return getattr(value, 'coordinate', value)


def _vopt_validated_cells(ws):
    cells = set()
    for dv in ws.data_validations.dataValidation:
        for rng in dv.cells.ranges:
            text = str(rng).replace('$', '')
            try:
                min_col, min_row, max_col, max_row = range_boundaries(text)
            except ValueError:
                cells.add(text)
                continue
            if max_row > 10000:
                max_row = max(min_row, 200)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cells.add(f"{get_column_letter(col)}{row}")
    return cells


class _VoptValidationRanges(list):
    def __init__(self, ws, ranges):
        super().__init__(ranges)
        self.ws = ws

    def __contains__(self, expected):
        expected = str(expected).replace('$', '')
        if super().__contains__(expected):
            return True
        try:
            min_col, min_row, max_col, max_row = range_boundaries(expected)
        except ValueError:
            return super().__contains__(expected)
        target = {f"{get_column_letter(col)}{row}" for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1)}
        return target.issubset(_vopt_validated_cells(self.ws))


def _vopt_dv_ranges(ws):
    ranges = []
    for dv in ws.data_validations.dataValidation:
        ranges.extend(str(rng).replace('$', '') for rng in dv.cells.ranges)
    return _VoptValidationRanges(ws, ranges)


def _vopt_ordered_subset(expected, actual):
    actual_iter = iter([_vopt_norm_text(item) for item in actual])
    for item in [_vopt_norm_text(value) for value in expected]:
        for candidate in actual_iter:
            if item == candidate:
                break
        else:
            return False
    return True



def _norm_ranges(ranges):
    return sorted(str(rng) for rng in ranges)


def _sheet_row_highlighted(ws, row_idx, min_col, max_col):
    for col in range(min_col, max_col + 1):
        rgb = cell_fill_rgb(ws.cell(row_idx, col))
        if rgb and rgb not in ('FFFFFF', '000000', '00000000'):
            return True
    return False


def _formula(value, label):
    assert isinstance(value, str) and value.strip().startswith('='), f"{label}: expected a live formula, found {value!r}"
    return _vopt_norm_formula(value)


def _assert_release_flag_formula(value, row):
    formula = _formula(value, f'Vendor Tracker!J{row}')
    assert f'F{row}' in formula or '$F' in formula or 'STATUS' in formula, (
        f'Vendor Tracker!J{row}: formula should depend on row status'
    )
    for token in ('PROVIDED', 'PARTIAL', 'MISSING', 'READY', 'WATCH', 'HOLD'):
        assert token in formula, f'Vendor Tracker!J{row}: formula missing release-flag token {token!r}'


def _assert_summary_formula(value, label, source_col, criterion):
    formula = _formula(value, f'Summary formula {label}')
    compact = formula.replace("'", '')
    assert 'VENDORTRACKER' in compact, f'Summary formula {label}: should reference Vendor Tracker'
    assert f'{source_col}2:{source_col}7' in compact or f'${source_col}$2:${source_col}$7' in compact, (
        f'Summary formula {label}: should summarize Vendor Tracker {source_col}2:{source_col}7'
    )
    assert criterion.upper() in compact, f'Summary formula {label}: missing criterion {criterion!r}'
    assert any(fn in compact for fn in ('COUNTIF', 'COUNTIFS', 'SUMPRODUCT')), (
        f'Summary formula {label}: expected a counting formula, found {value!r}'
    )


def _cf_fill_present(rule):
    fill = getattr(getattr(rule, 'dxf', None), 'fill', None)
    if fill is None:
        return False
    for color in (getattr(fill, 'fgColor', None), getattr(fill, 'bgColor', None), getattr(fill, 'start_color', None)):
        if color is None:
            continue
        rgb = getattr(color, 'rgb', None)
        indexed = getattr(color, 'indexed', None)
        theme = getattr(color, 'theme', None)
        if rgb or indexed is not None or theme is not None:
            return True
    return getattr(fill, 'patternType', None) is not None


def _range_covers_row(range_text, row_idx, min_col, max_col):
    refs = re.findall(r'\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?', range_text)
    for ref in refs:
        clean = ref.replace('$', '')
        if ':' not in clean:
            clean = f'{clean}:{clean}'
        try:
            c1, r1, c2, r2 = range_boundaries(clean)
        except ValueError:
            continue
        if r1 <= row_idx <= r2 and c1 <= min_col and c2 >= max_col:
            return True
    return False


def _conditional_row_highlighted(ws, row_idx, min_col, max_col):
    for rng in ws.conditional_formatting:
        if not _range_covers_row(str(rng), row_idx, min_col, max_col):
            continue
        for rule in ws.conditional_formatting[rng]:
            if _cf_fill_present(rule):
                return True
    return False


def _row_highlighted(ws, row_idx, min_col, max_col):
    return _sheet_row_highlighted(ws, row_idx, min_col, max_col) or _conditional_row_highlighted(
        ws, row_idx, min_col, max_col
    )


def _range_signatures(ws, refs):
    return [style_signature(ws[ref]) for ref in refs]


def test_output_exists_and_is_xlsx():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)


def test_required_cells_and_formulas():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == EXPECT['sheet_order'], f"Unexpected sheet order: {wb.sheetnames!r}"

    for ref, expected in EXPECT['exact_cells'].items():
        sheet_name, cell_ref = ref.split('!')
        actual = wb[sheet_name][cell_ref].value
        actual = '' if actual is None else str(actual)
        assert actual == str(expected), f"{ref}: expected {expected!r}, found {actual!r}"

    for row in (2, 3, 4, 6, 7):
        _assert_release_flag_formula(wb['Vendor Tracker'][f'J{row}'].value, row)

    summary_checks = {
        'B3': ('Provided evidence', 'F', 'Provided'),
        'B4': ('Partial evidence', 'F', 'Partial'),
        'B5': ('Missing evidence', 'F', 'Missing'),
        'B6': ('Ready vendors', 'J', 'Ready'),
        'B7': ('Watch vendors', 'J', 'Watch'),
        'B8': ('Hold vendors', 'J', 'Hold'),
    }
    for cell_ref, (label, source_col, criterion) in summary_checks.items():
        _assert_summary_formula(wb['Summary'][cell_ref].value, label, source_col, criterion)


def test_formula_results_are_consistent_when_cached():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    for ref, expected in EXPECT['formula_result_cells'].items():
        sheet_name, cell_ref = ref.split('!')
        ws = wb[sheet_name]
        row = int(cell_ref[1:])
        status = ws[f'F{row}'].value
        required = ws[f'D{row}'].value
        if required == 'Yes' and status == 'Provided':
            computed = 'Ready'
        elif required == 'Yes' and status == 'Partial':
            computed = 'Watch'
        elif required == 'Yes' and status == 'Missing':
            computed = 'Hold'
        else:
            computed = 'Ready'
        assert computed == expected, f"{ref}: expected computed value {expected!r}, found {computed!r}"


def test_preservation_predicates():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    in_wb = load_workbook(INPUT_PATH, data_only=False)

    for sheet_name in EXPECT['unchanged_sheets']:
        assert workbook_values_signature(wb)[sheet_name] == workbook_values_signature(in_wb)[sheet_name], (
            f'Sheet values changed unexpectedly: {sheet_name}'
        )

    for sheet_name in EXPECT['hidden_sheets']:
        assert wb[sheet_name].sheet_state in ('hidden', 'veryHidden'), f"{sheet_name} should remain hidden"

    defined_names = {dn.name for dn in wb.defined_names.values()}
    for name in EXPECT['defined_names']:
        assert name in defined_names, f"Missing defined name: {name}"

    for sheet_name, ranges in EXPECT['data_validation_ranges'].items():
        actual = _vopt_dv_ranges(wb[sheet_name])
        for expected in ranges:
            assert expected in actual, f"Missing data validation range {sheet_name}!{expected}; found {actual!r}"

    for key, expected_ref in EXPECT['table_refs'].items():
        sheet_name, table_name = key.split('!')
        actual_ref = _vopt_table_ref(wb[sheet_name], table_name)
        assert actual_ref == expected_ref, f"{key}: expected table ref {expected_ref}, found {actual_ref}"

    for sheet_name, pane in EXPECT['freeze_panes'].items():
        assert _vopt_freeze_pane(wb[sheet_name].freeze_panes) == pane, f"{sheet_name}: expected freeze pane {pane}"


def test_formatting_and_highlights():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    for spec in EXPECT['header_style_matches']:
        target = _range_signatures(wb[spec['target_sheet']], spec['target_cells'])
        reference = _range_signatures(wb[spec['reference_sheet']], spec['reference_cells'])
        assert target == reference, f"Header style mismatch for {spec['target_sheet']}"

    for spec in EXPECT['highlight_rows']:
        assert _row_highlighted(wb[spec['sheet']], spec['row'], spec['min_col'], spec['max_col']), (
            f"Expected highlighted row {spec['sheet']}!{spec['row']}"
        )
