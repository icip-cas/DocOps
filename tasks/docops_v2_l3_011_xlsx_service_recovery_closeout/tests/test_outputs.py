import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

sys.path.insert(0, str(Path(__file__).parent))
from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(__import__('os').environ.get('TASK_METADATA_PATH', '/tests/task_metadata.json'))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / 'task_metadata.json'
META = json.loads(META_PATH.read_text())
INPUT_PATH = Path(META['input_path'])
OUTPUT_PATH = Path(META['output_path'])

FRONT_SHEETS = ['Dashboard', 'Release Queue', 'Exception Review', 'Customer Notices']
REQUIRED_SHEETS = FRONT_SHEETS + [
    'Raw Intake',
    'SLA Policy',
    'Owner Map',
    'Part Catalog',
    'Technician Roster',
    'Scratch Notes',
    'Archive',
]

RQ_HEADERS = [
    'Ticket', 'Customer', 'Tier', 'Region', 'Opened', 'Closed', 'Status',
    'Issue Type', 'Part SKU', 'Priority', 'Owner', 'SLA Hours', 'Age Hours',
    'SLA Breach', 'Part Cost', 'Escalation Reason', 'Publish Note',
    'Customer Notice', 'Ready To Publish'
]

REVIEW_TS_FALLBACK = datetime(2026, 6, 5, 12, 0)
EXCEPTION_TICKETS = {'TCK-1001', 'TCK-1003', 'TCK-1004', 'TCK-1006', 'TCK-1008'}
NOTICE_TICKETS = {
    'TCK-1001', 'TCK-1002', 'TCK-1004', 'TCK-1005',
    'TCK-1006', 'TCK-1007', 'TCK-1008', 'TCK-1009'
}
FORBIDDEN_VISIBLE = ['billing dispute', 'priya@alden.example', 'VIP renewal at risk']


def _compact(value):
    return normalize_text(str(value or '')).replace(' ', '').replace('-', '')


def _headers(ws):
    return [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]


def _header_map(ws):
    return {norm_cell(value): idx + 1 for idx, value in enumerate(_headers(ws)) if norm_cell(value)}


def _require_col(ws, *names):
    headers = _header_map(ws)
    for name in names:
        if name in headers:
            return headers[name]
    raise AssertionError(f'{ws.title}: missing expected column, accepted names={names!r}, headers={list(headers)!r}')


def _direct_ref(value):
    if not isinstance(value, str):
        return None
    text = value.strip()
    match = re.fullmatch(r"=\s*'([^']+)'!\$?([A-Z]{1,3})\$?(\d+)", text, flags=re.I)
    if match:
        return match.group(1), match.group(2).upper(), int(match.group(3))
    match = re.fullmatch(r"=\s*([A-Za-z0-9_ ]+)!\$?([A-Z]{1,3})\$?(\d+)", text, flags=re.I)
    if match:
        return match.group(1).strip(), match.group(2).upper(), int(match.group(3))
    return None


def _display_value(wb, value, depth=0):
    ref = _direct_ref(value)
    if not ref or depth > 4:
        return value
    sheet, col, row = ref
    if sheet not in wb.sheetnames:
        return value
    return _display_value(wb, wb[sheet][f'{col}{row}'].value, depth + 1)


def _cell_value(wb, ws, row, col):
    return _display_value(wb, ws.cell(row, col).value)


def _truthy(value):
    text = normalize_text(str(value or ''))
    return value is True or text in ('true', 'yes', 'y', '1', 'ready')


def _falsey(value):
    text = normalize_text(str(value or ''))
    return value is False or text in ('false', 'no', 'n', '0', 'hold', '')


def _table_ref_contains(ws, name_part):
    name_part = name_part.lower()
    for name in ws.tables:
        if name_part in str(name).lower():
            return ws.tables[name].ref
    return table_ref(ws)


def _assert_table_has_rows(ws, name_part, min_rows):
    ref = _table_ref_contains(ws, name_part)
    _, min_row, _, max_row = range_boundaries(ref)
    assert min_row == 1 and max_row >= min_rows, f'{ws.title}: table {ref} does not cover required rows'
    return ref


def _sheet_text_values(wb, ws):
    values = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = _cell_value(wb, ws, row, col)
            if value is not None and not (isinstance(value, str) and value.startswith('=')):
                values.append(str(value))
    return '\n'.join(values)


def _row_by_ticket(wb, ws, ticket):
    ticket_col = _require_col(ws, 'Ticket')
    for row in range(2, ws.max_row + 1):
        if _cell_value(wb, ws, row, ticket_col) == ticket:
            return row
    raise AssertionError(f'{ws.title}: missing ticket {ticket}')


def _ticket_set(wb, ws):
    ticket_col = _require_col(ws, 'Ticket')
    values = set()
    for row in range(2, ws.max_row + 1):
        value = _cell_value(wb, ws, row, ticket_col)
        if value:
            values.add(str(value))
    return values


def _direct_row_fill(ws, row, min_col=1, max_col=19):
    for col in range(min_col, min(max_col, ws.max_column) + 1):
        rgb = (cell_fill_rgb(ws.cell(row, col)) or '').upper()
        if rgb not in ('', 'FFFFFF', '000000', '00000000'):
            return True
    return False


def _conditional_fill_covers(ws):
    ranges = [str(rng) for rng in ws.conditional_formatting]
    return any('A2' in rng and any(end in rng for end in ('S10', 'T10')) for rng in ranges)


def _row_highlighted(ws, row):
    return _direct_row_fill(ws, row) or _conditional_fill_covers(ws)


def _assert_release_sheets_do_not_expose_internal_notes(wb):
    for sheet_name in FRONT_SHEETS[1:]:
        ws = wb[sheet_name]
        headers = [_compact(h) for h in _headers(ws)]
        assert 'internalnote' not in headers, f'{sheet_name}: Internal Note must not be visible in release-facing sheets'
        visible_text = normalize_text(_sheet_text_values(wb, ws))
        for forbidden in FORBIDDEN_VISIBLE:
            assert normalize_text(forbidden) not in visible_text, (
                f'{sheet_name}: forbidden internal note leaked into visible release output: {forbidden}'
            )


def _map_from_sheet(ws, key_col, value_col):
    return {ws.cell(row, key_col).value: ws.cell(row, value_col).value for row in range(2, ws.max_row + 1)}


def _expected_priority(tier, issue_type):
    if tier == 'Platinum' or issue_type == 'Power':
        return 'P1'
    if tier == 'Gold':
        return 'P2'
    if tier == 'Silver':
        return 'P3'
    return 'P4'


def _review_timestamp(wb, raw):
    value = wb['Dashboard']['B2'].value
    assert isinstance(value, datetime), f'Dashboard!B2 should be a fixed review timestamp, found {value!r}'
    latest_opened = max(raw.cell(row, 5).value for row in range(2, raw.max_row + 1))
    assert value >= latest_opened, 'Dashboard!B2 should not predate the latest ticket opened time'
    return value


def _expected_age(opened, closed, review_ts):
    end = closed or review_ts
    return (end - opened).total_seconds() / 3600


def _assert_formula_or_value(wb, cell, expected, label, tokens=()):
    value = _display_value(wb, cell.value)
    if isinstance(cell.value, str) and cell.value.startswith('=') and _direct_ref(cell.value) is None:
        formula = _compact(normalize_formula(cell.value)).upper()
        for token in tokens:
            assert _compact(token).upper() in formula, f'{label}: formula missing token {token!r}: {cell.value!r}'
        return
    if isinstance(expected, float):
        assert abs(float(value) - expected) < 0.3, f'{label}: expected about {expected}, found {value!r}'
    else:
        assert value == expected, f'{label}: expected {expected!r}, found {value!r}'


def _check_dashboard(wb):
    dash = wb['Dashboard']
    title = normalize_text(str(dash['A1'].value or ''))
    assert 'service recovery' in title and 'dashboard' in title

    expected_metrics = {
        ('open', 'ticket'): ['COUNTIF', 'OPEN'],
        ('waiting', 'parts'): ['COUNTIF', 'WAITINGONPARTS'],
        ('sla', 'breach'): ['COUNTIF'],
        ('ready', 'notice'): ['COUNTIF'],
        ('part', 'exposure'): ['SUM'],
        ('west', 'breach'): ['COUNTIFS', 'WEST'],
    }
    for label_terms, formula_terms in expected_metrics.items():
        matched_formula = None
        for row in range(1, dash.max_row + 1):
            label = _compact(dash.cell(row, 1).value)
            if all(term in label for term in label_terms):
                matched_formula = dash.cell(row, 2).value
                break
        assert isinstance(matched_formula, str) and matched_formula.startswith('='), (
            f'Dashboard metric {label_terms}: expected a formula'
        )
        formula = _compact(normalize_formula(matched_formula)).upper()
        assert 'RELEASE' in formula or 'QUEUE' in formula, f'Dashboard metric {label_terms}: formula should use Release Queue'
        for term in formula_terms:
            assert term in formula, f'Dashboard metric {label_terms}: formula missing {term!r}: {matched_formula!r}'


def _check_release_queue(wb, src):
    rq = wb['Release Queue']
    raw = src['Raw Intake']
    review_ts = _review_timestamp(wb, raw)
    owners = _map_from_sheet(src['Owner Map'], 1, 2)
    sla_hours = _map_from_sheet(src['SLA Policy'], 1, 2)
    part_costs = _map_from_sheet(src['Part Catalog'], 1, 3)

    _assert_table_has_rows(rq, 'releasequeue', 10)
    headers = _header_map(rq)
    missing = [header for header in RQ_HEADERS if header not in headers]
    assert not missing, f'Release Queue missing required headers: {missing}'
    assert freeze_pane_ref(rq) == 'A2'
    assert expand_sqref_cells(sheet_data_validation_ranges(rq)) == expand_sqref_cells(['G2:G10'])

    raw_headers = _header_map(raw)
    for row in range(2, 11):
        ticket = raw.cell(row, raw_headers['Ticket']).value
        assert _cell_value(wb, rq, row, headers['Ticket']) == ticket
        for field in ['Customer', 'Tier', 'Region', 'Opened', 'Closed', 'Status', 'Issue Type', 'Part SKU']:
            actual = _cell_value(wb, rq, row, headers[field])
            expected = raw.cell(row, raw_headers[field]).value
            assert actual == expected, f'Release Queue {field} row {row}: expected {expected!r}, found {actual!r}'

        tier = raw.cell(row, raw_headers['Tier']).value
        region = raw.cell(row, raw_headers['Region']).value
        opened = raw.cell(row, raw_headers['Opened']).value
        closed = raw.cell(row, raw_headers['Closed']).value
        status = raw.cell(row, raw_headers['Status']).value
        issue = raw.cell(row, raw_headers['Issue Type']).value
        part = raw.cell(row, raw_headers['Part SKU']).value
        priority = _expected_priority(tier, issue)
        age = _expected_age(opened, closed, review_ts)
        breach = age > sla_hours[priority]

        _assert_formula_or_value(wb, rq.cell(row, headers['Priority']), priority, f'Priority row {row}', ('P1', 'P2', 'P3', 'P4'))
        _assert_formula_or_value(wb, rq.cell(row, headers['Owner']), owners[region], f'Owner row {row}', ('Owner Map', region))
        _assert_formula_or_value(wb, rq.cell(row, headers['SLA Hours']), sla_hours[priority], f'SLA Hours row {row}', ('SLA Policy', priority))
        _assert_formula_or_value(wb, rq.cell(row, headers['Age Hours']), age, f'Age Hours row {row}', ('Dashboard', 'B2'))
        _assert_formula_or_value(wb, rq.cell(row, headers['Part Cost']), part_costs[part], f'Part Cost row {row}', ('Part Catalog', part))

        breach_cell = rq.cell(row, headers['SLA Breach'])
        if isinstance(breach_cell.value, str) and breach_cell.value.startswith('='):
            formula = _compact(normalize_formula(breach_cell.value)).upper()
            assert '>' in formula or 'SLA' in formula, f'SLA Breach row {row}: formula should compare age to SLA'
        elif breach:
            assert _truthy(breach_cell.value), f'SLA Breach row {row}: expected true/yes'
        else:
            assert _falsey(breach_cell.value), f'SLA Breach row {row}: expected false/no'

        ready = rq.cell(row, headers['Ready To Publish'])
        if isinstance(ready.value, str) and ready.value.startswith('='):
            formula = _compact(normalize_formula(ready.value)).upper()
            assert 'WAITINGONPARTS' in formula and 'DONOTPUBLISH' in formula, (
                f'Ready To Publish row {row}: formula should block waiting-on-parts/do-not-publish rows'
            )
        elif ticket == 'TCK-1003':
            assert _falsey(ready.value), f'Ready To Publish row {row}: waiting-on-parts sensitive row should not publish'
        else:
            assert _truthy(ready.value), f'Ready To Publish row {row}: expected ready'

    for ticket in EXCEPTION_TICKETS:
        assert _row_highlighted(rq, _row_by_ticket(wb, rq, ticket)), f'Expected highlighted row for exception ticket {ticket}'
    for ticket in {'TCK-1002', 'TCK-1005', 'TCK-1007', 'TCK-1009'}:
        assert not _direct_row_fill(rq, _row_by_ticket(wb, rq, ticket)), f'Non-exception row should not have direct fill: {ticket}'


def _check_exception_review(wb):
    exc = wb['Exception Review']
    _assert_table_has_rows(exc, 'exceptionreview', 6)
    _require_col(exc, 'Ticket')
    _require_col(exc, 'Customer')
    _require_col(exc, 'Reason', 'Escalation Reason', 'Exception Reason', 'Review Flag')
    _require_col(exc, 'Owner')
    assert _ticket_set(wb, exc) == EXCEPTION_TICKETS, f'Exception Review tickets should be {sorted(EXCEPTION_TICKETS)}'
    text = normalize_text(_sheet_text_values(wb, exc))
    assert 'part' in text or any('WAITINGONPARTS' in _compact(cell.value).upper() for row in exc.iter_rows() for cell in row), (
        'Exception Review must identify the waiting-on-parts exception'
    )


def _check_customer_notices(wb):
    notices = wb['Customer Notices']
    _assert_table_has_rows(notices, 'customernotices', 9)
    _require_col(notices, 'Ticket')
    _require_col(notices, 'Customer')
    _require_col(notices, 'Owner')
    _require_col(notices, 'Customer Notice', 'Notice')
    tickets = _ticket_set(wb, notices)
    assert tickets == NOTICE_TICKETS, f'Customer Notices tickets should be {sorted(NOTICE_TICKETS)}, found {sorted(tickets)}'
    assert 'TCK-1003' not in tickets, 'Waiting-on-parts sensitive row should not be in Customer Notices'

    customer_col = _require_col(notices, 'Customer')
    owner_col = _require_col(notices, 'Owner')
    notice_col = _require_col(notices, 'Customer Notice', 'Notice')
    for ticket in NOTICE_TICKETS:
        row = _row_by_ticket(wb, notices, ticket)
        notice = normalize_text(str(_cell_value(wb, notices, row, notice_col) or ''))
        if notice:
            assert normalize_text(str(_cell_value(wb, notices, row, customer_col))) in notice
            assert normalize_text(str(_cell_value(wb, notices, row, owner_col))) in notice


def test_service_recovery_closeout_longflow_workbook():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)
    wb = load_workbook(OUTPUT_PATH)
    src = load_workbook(INPUT_PATH)

    assert wb.sheetnames[:4] == FRONT_SHEETS, f'Release-facing sheets should come first: {FRONT_SHEETS}'
    for sheet in REQUIRED_SHEETS:
        assert sheet in wb.sheetnames, f'Missing required sheet: {sheet}'
    assert 'Legacy Dashboard' not in wb.sheetnames

    for sheet in ['Raw Intake', 'SLA Policy', 'Owner Map', 'Part Catalog', 'Technician Roster']:
        assert workbook_values_signature(wb)[sheet] == workbook_values_signature(src)[sheet], (
            f'Source sheet values changed unexpectedly: {sheet}'
        )
    for sheet in ['Raw Intake', 'Scratch Notes', 'Archive']:
        assert sheet_hidden_state_ok(wb[sheet], 'hidden'), f'{sheet} should be hidden for release'

    _assert_release_sheets_do_not_expose_internal_notes(wb)
    _check_dashboard(wb)
    _check_release_queue(wb, src)
    _check_exception_review(wb)
    _check_customer_notices(wb)
