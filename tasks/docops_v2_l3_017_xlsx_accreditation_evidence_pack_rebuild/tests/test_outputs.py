import json
import os
from pathlib import Path

from openpyxl import load_workbook

from verifier_utils import cell_fill_rgb, cell_font_rgb, normalize_formula, normalize_text, run_preflight

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]
INPUT_PATH = Path(os.environ.get("INPUT_PATH", META["input_path"]))
OUTPUT_PATH = Path(os.environ.get("OUTPUT_PATH", META["output_path"]))


def sheet_text(wb):
    parts = list(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
    return "\n".join(parts)


def row_values(ws, row_idx, max_col):
    return [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]


def sheet_values_text(ws):
    parts = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                parts.append(str(cell.value))
    return normalize_text("\n".join(parts))


def formulas(ws):
    values = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                values.append(cell.value)
    return values


def has_date_value(cell, y, m, d):
    value = cell.value
    return (
        getattr(value, "year", None) == y
        and getattr(value, "month", None) == m
        and getattr(value, "day", None) == d
    ) or str(value)[:10] == f"{y:04d}-{m:02d}-{d:02d}"


def assert_table(ws, expected_rows):
    max_col = len(expected_rows[0])
    assert row_values(ws, 1, max_col) == expected_rows[0]
    assert ws.max_row == len(expected_rows), f"{ws.title}: wrong row count"
    for r_idx, expected in enumerate(expected_rows[1:], start=2):
        actual = row_values(ws, r_idx, max_col)
        assert actual == expected, f"{ws.title} row {r_idx}: expected {expected!r}, found {actual!r}"


def tab_color(ws):
    color = ws.sheet_properties.tabColor
    if color is None or color.rgb is None:
        return None
    rgb = color.rgb.upper()
    return rgb[-6:] if len(rgb) >= 6 else rgb


def assert_summary(ws):
    text = sheet_values_text(ws)
    summary_formulas = formulas(ws)
    assert "accreditation evidence package" in text or "metric" in text
    assert len(summary_formulas) >= 4
    joined = normalize_formula("\n".join(summary_formulas))
    for sheet in ["Evidence Register", "Finding Tracker"]:
        token = sheet.replace(" ", "").upper()
        assert token in joined, f"Executive Summary formulas should reference {sheet}"


def assert_header_style(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(1, col)
        if cell.value is None:
            break
        assert cell_fill_rgb(cell) == EXPECT["header_fill"], f"{ws.title}!{cell.coordinate}: wrong header fill"
        assert cell_font_rgb(cell) == EXPECT["header_font"], f"{ws.title}!{cell.coordinate}: wrong header font"
        assert cell.font.bold, f"{ws.title}!{cell.coordinate}: header not bold"
    assert ws["A2"].font.name == EXPECT["body_font"], f"{ws.title}: body font should be Aptos"


def row_by_key(ws, key):
    for row in ws.iter_rows(min_row=2):
        if row[0].value == key:
            return [cell.value for cell in row]
    raise AssertionError(f"{ws.title}: missing row keyed {key!r}")


def assert_contains_terms(ws, terms, label):
    text = sheet_values_text(ws)
    missing = [term for term in terms if normalize_text(term) not in text]
    assert not missing, f"{label}: missing {missing}"


def test_output_exists_and_is_xlsx():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)
    assert OUTPUT_PATH != INPUT_PATH


def test_sheets_tab_colors_and_private_material_removed():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    for sheet in EXPECT["sheet_order"]:
        assert tab_color(wb[sheet]) is not None, f"{sheet}: missing public tab color"
    text = normalize_text(sheet_text(wb))
    hits = [p for p in EXPECT["forbidden_phrases"] if normalize_text(p) in text]
    assert not hits, f"Forbidden phrases still present: {hits}"


def test_public_tables_and_formulas():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert_summary(wb["Executive Summary"])
    evidence = wb["Evidence Register"]
    assert row_values(evidence, 1, 8) == EXPECT["evidence_rows"][0]
    for ev_id in ["EV-001", "EV-002", "EV-003", "EV-004", "EV-012"]:
        row_by_key(evidence, ev_id)
    assert row_by_key(evidence, "EV-001")[3] == "Dr. Mina Zhou"
    assert row_by_key(evidence, "EV-003")[1] == "Standard 2.B.1"
    assert row_by_key(evidence, "EV-012")[2] == "Accreditation response calendar"
    for row_idx in range(2, evidence.max_row + 1):
        if evidence.cell(row_idx, 1).value:
            assert has_date_value(evidence.cell(row_idx, 6), 2026, 8, 15)

    standards = wb["Standards Matrix"]
    assert "Standard" in row_values(standards, 1, standards.max_column)[0]
    assert_contains_terms(standards, ["Standard 1.A", "Standard 1.B", "Standard 2.B.1"], "Standards Matrix")
    standard_formulas = formulas(standards)
    assert standard_formulas, "Standards Matrix should contain formulas"
    assert "EVIDENCEREGISTER" in normalize_formula("\n".join(standard_formulas))

    findings = wb["Finding Tracker"]
    assert_contains_terms(findings, ["F-001", "Standard 2.B.1", "Priya Raman", "Open"], "Finding Tracker")

    workplan = wb["Owner Workplan"]
    workplan_formulas = formulas(workplan)
    assert workplan_formulas, "Owner Workplan should contain formulas"

    risk = wb["Risk Heatmap"]
    assert row_values(risk, 1, 5) == EXPECT["risk_rows"][0]
    assert row_by_key(risk, "R-001")[2:4] == [4, "High"]
    assert row_by_key(risk, "R-002")[2:4] == [3, "Medium"]

    cleanup = wb["Citation Cleanup Log"]
    for term in ["Standard 2.B.1", "Accreditation", "Dr. Mina Zhou", "2026-08-15"]:
        assert_contains_terms(cleanup, [term], "Citation Cleanup Log")
    assert "risk" in sheet_values_text(cleanup) and "4" in sheet_values_text(cleanup)
    assert "yes" in sheet_values_text(cleanup) or "applied" in sheet_values_text(cleanup)

    style = wb["Publication Style Guide"]
    assert_contains_terms(style, ["1B4D3E", "Aptos"], "Publication Style Guide")
    assert "white" in sheet_values_text(style) and "bold" in sheet_values_text(style)
    assert "risk" in sheet_values_text(style) or "heatmap" in sheet_values_text(style)

    appendix = wb["Appendix Index"]
    assert "evidence" in sheet_values_text(appendix)
    assert "appendix" in sheet_values_text(appendix) or "sheet" in sheet_values_text(appendix)


def test_structure_and_number_formats():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    for sheet in EXPECT["sheet_order"][1:]:
        assert wb[sheet].freeze_panes == "A2"
        assert wb[sheet].auto_filter.ref is not None, f"{sheet}: missing autofilter"
    for row in range(2, wb["Evidence Register"].max_row + 1):
        if wb["Evidence Register"][f"A{row}"].value:
            assert wb["Evidence Register"][f"F{row}"].number_format.lower() == EXPECT["date_format"]
            assert wb["Evidence Register"][f"G{row}"].number_format in {EXPECT["percent_format"], "0.0%"}
    for row in range(2, wb["Owner Workplan"].max_row + 1):
        for col in range(1, wb["Owner Workplan"].max_column + 1):
            cell = wb["Owner Workplan"].cell(row, col)
            if has_date_value(cell, 2026, 8, 15) or getattr(cell.value, "year", None) == 2026:
                assert cell.number_format.lower() == EXPECT["date_format"]


def test_style_and_heatmap_applied():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert cell_fill_rgb(wb["Executive Summary"]["A1"]) == EXPECT["summary_title_fill"]
    for sheet, max_col in {
        "Evidence Register": 8,
        "Standards Matrix": 5,
        "Finding Tracker": 6,
        "Owner Workplan": 8,
        "Risk Heatmap": 5,
        "Citation Cleanup Log": 5,
        "Publication Style Guide": 2,
        "Appendix Index": 4
    }.items():
        assert_header_style(wb[sheet], max_col)
    fills = EXPECT["risk_level_fills"]
    for row in range(2, wb["Risk Heatmap"].max_row + 1):
        level = wb["Risk Heatmap"][f"D{row}"].value
        if level in fills:
            assert cell_fill_rgb(wb["Risk Heatmap"][f"D{row}"]) == fills[level], f"Risk row {row}: wrong heatmap fill"


def test_source_artifact_was_not_modified():
    source = load_workbook(INPUT_PATH, data_only=False)
    assert "Evidence_DRAFT_PRIVATE" in source.sheetnames
    assert "legal hold" in source.sheetnames
    text = normalize_text(sheet_text(source))
    assert "board-room dispute" in text
    assert "reviewer personal phone" in text
