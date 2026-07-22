import json
import os
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from pypdf import PdfReader

from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]


def _norm(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value).replace("\u2013", "-").replace("\u2014", "-").replace("\xa0", " ")).strip()


def _norm_formula(value):
    return re.sub(r"\s+", "", _norm(value)).upper()


def _norm_key(value):
    return re.sub(r"[^a-z0-9]+", "", _norm(value).casefold())


def _path(kind):
    env = {"pdf": "PDF_OUTPUT_PATH", "xlsx": "XLSX_OUTPUT_PATH"}[kind]
    key = {"pdf": "pdf_output", "xlsx": "xlsx_output"}[kind]
    return Path(os.environ.get(env, EXPECT[key]))


def _pdf_pages(path):
    reader = PdfReader(str(path))
    return [page.extract_text() or "" for page in reader.pages]


def _rows(ws, start, end, max_col):
    return [[_norm(ws.cell(r, c).value) for c in range(1, max_col + 1)] for r in range(start, end + 1)]


def _expected(rows):
    return [[_norm(c) for c in row] for row in rows]


def _sheet(wb, expected_name):
    _assert_required_sheets(wb, [expected_name])
    return wb[expected_name]


def _table(ws, expected_name):
    for name in ws.tables.keys():
        if str(name).lower() == str(expected_name).lower():
            return ws.tables[name]
    raise AssertionError(f"Missing table {expected_name!r} on {ws.title!r}")


def _table_ref(ws, expected_name):
    return _table(ws, expected_name).ref


def _table_bounds(table):
    return range_boundaries(table.ref.replace("$", ""))


def _table_headers(ws, table):
    min_col, min_row, max_col, _ = _table_bounds(table)
    return [_norm(ws.cell(min_row, col).value) for col in range(min_col, max_col + 1)]


def _table_header_map(ws, table):
    min_col, _, _, _ = _table_bounds(table)
    return {_norm_key(header): idx for idx, header in enumerate(_table_headers(ws, table), start=min_col)}


def _require_headers(ws, table, expected_headers):
    actual = {_norm_key(header) for header in _table_headers(ws, table)}
    missing = [header for header in expected_headers if _norm_key(header) not in actual]
    assert not missing, f"{ws.title}: table {table.name} missing headers {missing}; found {_table_headers(ws, table)}"


def _table_rows(ws, table):
    _, min_row, _, max_row = _table_bounds(table)
    header_map = _table_header_map(ws, table)
    rows = []
    for row_idx in range(min_row + 1, max_row + 1):
        row = {"__row_idx": row_idx}
        for key, col_idx in header_map.items():
            row[key] = ws.cell(row_idx, col_idx)
        if any(cell.value is not None for key, cell in row.items() if key != "__row_idx"):
            rows.append(row)
    return rows


def _cell_by_header(row, header):
    key = _norm_key(header)
    assert key in row, f"Missing header {header!r}"
    return row[key]


def _value_matches(actual, expected):
    actual_text = _norm(actual)
    expected_text = _norm(expected)
    if actual_text.casefold() == expected_text.casefold():
        return True
    if len(expected_text) >= 12 and expected_text.casefold() in actual_text.casefold():
        return True
    return False


def _find_row(rows, key_header, key_value):
    for row in rows:
        if _value_matches(_cell_by_header(row, key_header).value, key_value):
            return row
    raise AssertionError(f"Could not find row with {key_header!r}={key_value!r}")


def _assert_row_contains(rows, headers, expected_row, key_header=None, skip_headers=()):
    key_header = key_header or headers[0]
    row = _find_row(rows, key_header, expected_row[headers.index(key_header)])
    for header, expected_value in zip(headers, expected_row):
        if header in skip_headers:
            continue
        actual = _cell_by_header(row, header).value
        assert _value_matches(actual, expected_value), f"{header}: expected {expected_value!r}, found {actual!r}"
    return row


def _range_covered(ws, expected):
    min_col, min_row, max_col, max_row = range_boundaries(expected.replace("$", ""))
    target = {f"{get_column_letter(col)}{row}" for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1)}
    cells = set()
    for dv in ws.data_validations.dataValidation:
        for rng in dv.cells.ranges:
            min_c, min_r, max_c, max_r = range_boundaries(str(rng).replace("$", ""))
            cells.update(f"{get_column_letter(col)}{row}" for row in range(min_r, max_r + 1) for col in range(min_c, max_c + 1))
    return target.issubset(cells)


def _clean_area(value):
    if isinstance(value, (list, tuple)):
        value = ",".join(str(item) for item in value)
    text = _norm(value).replace("'", "").replace("$", "").replace(" ", "")
    if "!" in text:
        text = text.split("!", 1)[1]
    return text.upper()


def _sheet_text(wb, sheets):
    _assert_required_sheets(wb, sheets)
    parts = []
    for sheet in sheets:
        for row in wb[sheet].iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
    return "\n".join(parts)


def _assert_required_sheets(wb, sheets):
    missing = [sheet for sheet in sheets if sheet not in wb.sheetnames]
    assert not missing, f"missing required workbook sheets: {missing}; found {wb.sheetnames}"


def _formula_text(wb, ref):
    sheet, cell = ref.split("!")
    value = wb[sheet][cell].value
    assert isinstance(value, str) and value.startswith("="), f"{ref} must contain a formula, found {value!r}"
    return _norm_formula(value).replace("$", "").replace("'", "")


def _summary_formula_texts(wb):
    ws = _sheet(wb, "Fulfillment Summary")
    texts = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                texts.append(_norm_formula(cell.value).replace("$", "").replace("'", ""))
    return texts


def _formula_mentions_range_or_table(text, sheet, cell_range, table_name, field_name=None):
    sheet_token = sheet.upper().replace(" ", "")
    range_token = cell_range.upper()
    table_token = table_name.upper()
    alt_table_token = re.sub(r"^TBL", "", table_token)
    if f"{sheet_token}!{range_token}" in text or f"{sheet_token}{range_token}" in text:
        return True
    if (table_token in text or alt_table_token in text) and (field_name is None or field_name.upper().replace(" ", "") in text):
        return True
    return False


def _assert_formula_semantic(wb, ref):
    summary_formulas = _summary_formula_texts(wb)
    if ref == "Fulfillment Summary!B4":
        assert any(
            _formula_mentions_range_or_table(text, "Merge Control", "B2:B6", "tblMergeControl", "Name")
            and any(func in text for func in ("COUNTA", "COUNT", "ROWS"))
            for text in summary_formulas
        ), f"{ref}: expected a participant-count formula"
    elif ref == "Fulfillment Summary!B5":
        assert any(
            _formula_mentions_range_or_table(text, "Packet QA", "A2:A11", "tblPacketQA", "Page")
            and any(func in text for func in ("COUNTA", "COUNT", "ROWS"))
            for text in summary_formulas
        ), f"{ref}: expected a packet-page-count formula"
    elif ref == "Fulfillment Summary!B6":
        assert any(
            "READY" in text and _formula_mentions_range_or_table(text, "Merge Control", "G2:G6", "tblMergeControl", "Status")
            for text in summary_formulas
        ), f"{ref}: formula must count Ready rows"
    elif ref == "Fulfillment Summary!B8":
        text = next((item for item in summary_formulas if "READY" in item and "HOLD" in item), "")
        assert "READY" in text and "HOLD" in text, f"{ref}: formula must return Ready/Hold"
        assert "AND" in text and all(token in text for token in ("=5", "=10", "=4")), f"{ref}: formula must combine the summary control checks"
    else:
        raise AssertionError(f"Unexpected formula check requested for {ref}")


def _defined_name_names(wb):
    try:
        values = wb.defined_names.values()
    except AttributeError:
        values = getattr(wb.defined_names, "definedName", [])
    return {dn.name for dn in values if getattr(dn, "name", None)}


def _assert_summary_rows(wb):
    ws = _sheet(wb, "Fulfillment Summary")
    table = _table(ws, "tblFulfillmentSummary")
    rows = _table_rows(ws, table)
    summary_text = _sheet_text(wb, ["Fulfillment Summary"])
    assert normalize_text("Community Repair Skills Weekend") in normalize_text(summary_text)
    formula_values = [_norm_formula(_cell_by_header(row, "Value").value) for row in rows if isinstance(_cell_by_header(row, "Value").value, str) and str(_cell_by_header(row, "Value").value).startswith("=")]
    assert any("MERGECONTROL" in formula or "TBLMERGECONTROL" in formula for formula in formula_values)
    assert any("PACKETQA" in formula or "TBLPACKETQA" in formula for formula in formula_values)
    assert any("READY" in formula and "HOLD" in formula for formula in formula_values)


def _expected_key_for_sheet(sheet):
    return {
        "Fulfillment Summary": "summary_rows",
        "Merge Control": "merge_rows",
        "Packet QA": "packet_rows",
        "Style QA": "style_rows",
        "Private Exceptions": "private_rows",
    }[sheet]


def test_outputs_exist():
    assert _path("pdf").exists()
    assert _path("xlsx").exists()
    assert _path("pdf").suffix.lower() == ".pdf"
    assert _path("xlsx").suffix.lower() == ".xlsx"


def test_pdf_packet_content_order_privacy():
    pages = _pdf_pages(_path("pdf"))
    assert len(pages) == EXPECT["pdf_page_count"]
    text = "\n".join(pages)
    require_all(text, EXPECT["pdf_required"], "merged pdf packet")
    forbid_any(text, EXPECT["pdf_forbidden"], "merged pdf packet")
    for row in EXPECT["packet_rows"][1:]:
        page_no, doc_type, name, expected_text, _ = row
        page = pages[int(page_no) - 1]
        page_norm = normalize_text(page)
        assert normalize_text(name) in page_norm, f"page {page_no}: missing name {name!r}"
        assert normalize_text(expected_text) in page_norm, f"page {page_no}: missing expected text {expected_text!r}"
        assert normalize_text(doc_type) in page_norm, f"page {page_no}: missing document type {doc_type!r}"
        if doc_type == "Certificate":
            assert normalize_text("Certificate of Completion") in page_norm
        if doc_type == "Badge":
            assert normalize_text("Badge") in page_norm


def test_workbook_structure_formulas_controls():
    wb = load_workbook(_path("xlsx"), data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    _assert_required_sheets(wb, EXPECT["sheet_order"])
    for sheet, (table_name, ref) in EXPECT["tables"].items():
        ws = _sheet(wb, sheet)
        table = _table(ws, table_name)
        assert table.ref == ref
        _require_headers(ws, table, EXPECT[_expected_key_for_sheet(sheet)][0])
    for sheet in EXPECT["hidden_sheets"]:
        assert _sheet(wb, sheet).sheet_state in ("hidden", "veryHidden")
    names = _defined_name_names(wb)
    for name in EXPECT["defined_names"]:
        assert name in names
    for ref in EXPECT["formula_cells"]:
        _assert_formula_semantic(wb, ref)
    for sheet, ranges in EXPECT["data_validation"].items():
        for rng in ranges:
            assert _range_covered(_sheet(wb, sheet), rng), f"{sheet}: missing validation over {rng}"
    for sheet, area in EXPECT["print_areas"].items():
        actual = _clean_area(_sheet(wb, sheet).print_area)
        target = _clean_area(area)
        assert target in actual or actual in target


def test_workbook_values_privacy():
    wb = load_workbook(_path("xlsx"), data_only=False)
    _assert_required_sheets(wb, EXPECT["sheet_order"])
    _assert_summary_rows(wb)
    for sheet, table_name, expected_rows, key_header in [
        ("Merge Control", "tblMergeControl", EXPECT["merge_rows"], "Seq"),
        ("Packet QA", "tblPacketQA", EXPECT["packet_rows"], "Page"),
    ]:
        ws = _sheet(wb, sheet)
        rows = _table_rows(ws, _table(ws, table_name))
        for expected_row in expected_rows[1:]:
            _assert_row_contains(rows, expected_rows[0], expected_row, key_header=key_header)
    style_text = normalize_text(_sheet_text(wb, ["Style QA"]))
    for token in ("1f4e5f", "e8b44b", "certificate", "badge"):
        assert token in style_text
    style_rows = _table_rows(_sheet(wb, "Style QA"), _table(_sheet(wb, "Style QA"), "tblStyleQA"))
    assert sum(1 for row in style_rows if _value_matches(_cell_by_header(row, "Status").value, "Pass")) >= 5
    private_text = _sheet_text(wb, ["Private Exceptions"])
    for phrase in EXPECT["pdf_forbidden"]:
        if phrase in {"private donor note", "discount dispute", "internal exception"}:
            continue
        assert normalize_text(phrase) in normalize_text(private_text)
    public = _sheet_text(wb, ["Fulfillment Summary", "Merge Control", "Packet QA", "Style QA"])
    forbid_any(public, EXPECT["pdf_forbidden"], "public workbook sheets")


def test_cross_output_consistency():
    pdf_text = "\n".join(_pdf_pages(_path("pdf")))
    wb = load_workbook(_path("xlsx"), data_only=False)
    _assert_required_sheets(wb, ["Fulfillment Summary", "Merge Control", "Packet QA"])
    public = _sheet_text(wb, ["Merge Control", "Packet QA"])
    for _, name, workshop, cert_id, _, note, _ in EXPECT["merge_rows"][1:]:
        for anchor in [name, workshop, cert_id, note]:
            assert normalize_text(anchor) in normalize_text(pdf_text)
            assert normalize_text(anchor) in normalize_text(public)
    _assert_formula_semantic(wb, "Fulfillment Summary!B8")
