import json
import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from verifier_utils import cell_fill_rgb, cell_font_rgb, normalize_formula, normalize_text, run_preflight

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]
INPUT_PATH = Path(os.environ.get("INPUT_PATH", META["input_path"]))
OUTPUT_PATH = Path(os.environ.get("OUTPUT_PATH", META["output_path"]))


def sheet_text(wb):
    parts = list(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
    return "\n".join(parts)


def values(ws, row, cols):
    return [ws.cell(row, c).value for c in range(1, cols + 1)]


def sheet_norm(ws):
    return normalize_text("\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None))


def workbook_norm(wb):
    return normalize_text(sheet_text(wb))


def assert_terms(text, terms, label):
    missing = [term for term in terms if normalize_text(term) not in text]
    assert not missing, f"{label}: missing semantic terms {missing}"


def assert_any(text, terms, label):
    assert any(normalize_text(term) in text for term in terms), f"{label}: missing one of {terms}"


def date_like_values(wb):
    vals = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, datetime):
                    vals.append(cell.value.strftime("%Y-%m-%d"))
                elif cell.value is not None:
                    vals.append(str(cell.value))
    return "\n".join(vals)


def formula_cells(ws):
    return [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]


def tab_color(ws):
    c = ws.sheet_properties.tabColor
    return None if c is None or c.rgb is None else c.rgb.upper()[-6:]


def assert_summary(ws):
    text = sheet_norm(ws)
    assert "governance" in text or "ai" in text or "metric" in text
    formulas = formula_cells(ws)
    assert len(formulas) >= 3, "Governance Summary should use formulas for public metrics"
    normalized = " ".join(normalize_formula(f) for f in formulas)
    assert any(fn in normalized for fn in ["COUNTA", "ROWS"]), "Governance Summary missing system-count formula"
    assert "COUNTIF" in normalized, "Governance Summary missing risk or incident count formula"
    assert "AVERAGE" in normalized, "Governance Summary missing average evaluation coverage formula"


def assert_header_style(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(1, col)
        assert cell_fill_rgb(cell) == EXPECT["header_fill"]
        assert cell_font_rgb(cell) == EXPECT["header_font"]
        assert cell.font.bold is not False


def assert_public_header_style(ws):
    cells = [ws.cell(1, col) for col in range(1, ws.max_column + 1) if ws.cell(1, col).value is not None]
    assert len(cells) >= 2, f"{ws.title}: missing public table header row"
    for cell in cells:
        assert cell_fill_rgb(cell) == EXPECT["header_fill"], f"{ws.title}!{cell.coordinate}: wrong header fill"
        assert cell_font_rgb(cell) == EXPECT["header_font"], f"{ws.title}!{cell.coordinate}: wrong header font"
        assert cell.font.bold is not False, f"{ws.title}!{cell.coordinate}: header not bold"


def assert_body_font(ws):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is not None and cell.font.name is not None:
                assert cell.font.name == EXPECT["body_font"], f"{ws.title}!{cell.coordinate}: body font should be Aptos"
                return


def assert_semantic_sheet(ws, required_terms, label, any_terms=()):
    text = sheet_norm(ws)
    assert_terms(text, required_terms, label)
    if any_terms:
        assert_any(text, any_terms, label)


def assert_column_formats(ws):
    headers = {col: normalize_text(str(ws.cell(1, col).value or "")) for col in range(1, ws.max_column + 1)}
    for row in range(2, ws.max_row + 1):
        for col, header in headers.items():
            cell = ws.cell(row, col)
            if cell.value is None:
                continue
            is_date_col = (
                header not in {"date format", "percent format"}
                and (
                    re.search(r"\bdate\b", header) is not None
                    or re.search(r"\bupdated\b", header) is not None
                    or re.search(r"\breviewed\b", header) is not None
                )
            )
            is_percent_col = header != "percent format" and any(term in header for term in ["coverage", "percent", "rate", "%"])
            if is_date_col:
                assert cell.number_format == EXPECT["date_format"], f"{ws.title}!{cell.coordinate}: wrong date format"
            if is_percent_col:
                assert cell.number_format == EXPECT["percent_format"], f"{ws.title}!{cell.coordinate}: wrong percent format"


def test_output_exists_and_is_xlsx():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)
    assert OUTPUT_PATH != INPUT_PATH


def test_sheets_tab_colors_and_private_material_removed():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    for sheet in EXPECT["sheet_order"]:
        assert tab_color(wb[sheet]), f"{sheet}: missing public sheet tab color"
    text = workbook_norm(wb)
    forbidden = EXPECT["forbidden_phrases"] + ["Goven", "Mappping", "Chatbot Triage v0.8", "2026-05-01", "Ravi Shah"]
    hits = [p for p in forbidden if normalize_text(p) in text]
    assert not hits, f"Forbidden phrases still present: {hits}"
    assert_terms(text, ["Govern", "Chatbot Triage v1.0", "Ravi Sharma"], "correction facts")
    assert "mapping" in text or "map" in text
    assert "2026-06-10" in date_like_values(wb)


def test_public_tables_and_summary_formulas():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert_summary(wb["Governance Summary"])
    assert_semantic_sheet(wb["AI System Inventory"], ["Chatbot Triage v1.0", "Ravi Sharma"], "AI System Inventory", ["Permit Summary Assistant", "Benefits"])
    assert_semantic_sheet(wb["RMF Function Matrix"], ["Govern"], "RMF Function Matrix", ["Mapping", "Map", "Measure", "Manage"])
    assert_semantic_sheet(wb["Risk Assessment Register"], ["High", "Medium", "Low"], "Risk Assessment Register", ["mitigation", "risk"])
    assert_semantic_sheet(wb["Evaluation Evidence"], ["2026-06-10"], "Evaluation Evidence", ["Bias", "Coverage", "Evaluation"])
    assert_semantic_sheet(wb["Incident Log"], [], "Incident Log", ["Incident", "Closed", "Open", "Status"])
    assert_semantic_sheet(wb["Human Oversight Plan"], [], "Human Oversight Plan", ["Human", "Review", "Reviewer", "Oversight"])
    assert_semantic_sheet(wb["Public Transparency Register"], [], "Public Transparency Register", ["Transparency", "Public", "Disclosure", "Notice"])
    assert_semantic_sheet(wb["Publication Style Guide"], ["243B53", "Aptos"], "Publication Style Guide", ["High C00000", "Risk heatmap", "tab"])
    assert_semantic_sheet(wb["Appendix Index"], [], "Appendix Index", ["Appendix", "Inventory", "Evidence", "Transparency"])


def test_structure_styles_and_formats():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    for sheet in EXPECT["sheet_order"][1:]:
        ws = wb[sheet]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref, f"{sheet}: missing autofilter"
        assert_public_header_style(ws)
        assert_body_font(ws)
        assert_column_formats(ws)


def test_risk_heatmap_applied():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    fills = EXPECT["risk_level_fills"]
    ws = wb["Risk Assessment Register"]
    found = {level: False for level in fills}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in fills and cell_fill_rgb(cell) == fills[cell.value]:
                found[cell.value] = True
    missing = [level for level, ok in found.items() if not ok]
    assert not missing, f"Risk Assessment Register missing heatmap fills for {missing}"


def test_source_artifact_was_not_modified():
    source = load_workbook(INPUT_PATH, data_only=False)
    assert "AI_Risk_DRAFT_PRIVATE" in source.sheetnames
    assert "legal hold" in source.sheetnames
    text = normalize_text(sheet_text(source))
    assert "prompt injection transcript" in text
    assert "personal phone" in text
