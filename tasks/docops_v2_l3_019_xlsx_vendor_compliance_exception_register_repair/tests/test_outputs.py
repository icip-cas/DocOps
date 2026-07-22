import json
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart
from openpyxl.utils import range_boundaries

from verifier_utils import cell_fill_rgb, cell_font_rgb, normalize_formula, normalize_text, run_preflight

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]
INPUT_PATH = Path(os.environ.get("INPUT_PATH", META["input_path"]))
OUTPUT_PATH = Path(os.environ.get("OUTPUT_PATH", META["output_path"]))


def sheet_text(wb):
    worksheets = [wb] if hasattr(wb, "iter_rows") and not hasattr(wb, "worksheets") else wb.worksheets
    parts = [] if len(worksheets) == 1 else list(wb.sheetnames)
    for ws in worksheets:
        parts.append(ws.title)
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
                if cell.comment is not None:
                    parts.append(cell.comment.text)
    return "\n".join(parts)


def tab_color(ws):
    color = ws.sheet_properties.tabColor
    return None if color is None or color.rgb is None else color.rgb.upper()[-6:]


def values_match(actual, expected):
    if expected == "" and actual is None:
        return True
    if isinstance(expected, str) and expected.startswith("="):
        return normalize_formula(actual) == normalize_formula(expected)
    return actual == expected


def assert_table_values(ws, expected_rows):
    assert ws.max_row == len(expected_rows)
    for row_idx, expected_row in enumerate(expected_rows, 1):
        actual = [ws.cell(row_idx, col).value for col in range(1, len(expected_row) + 1)]
        mismatches = [
            (col_idx, a, e)
            for col_idx, (a, e) in enumerate(zip(actual, expected_row), 1)
            if not values_match(a, e)
        ]
        assert not mismatches, f"{ws.title} row {row_idx} mismatches: {mismatches}"


def row_values(ws, row=1):
    return [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]


def normalized_headers(ws):
    return [normalize_text(str(value or "")) for value in row_values(ws, 1)]


def find_header_col(ws, required_terms, optional_terms=()):
    headers = normalized_headers(ws)
    for idx, header in enumerate(headers, 1):
        if all(term in header for term in required_terms) and (
            not optional_terms or any(term in header for term in optional_terms)
        ):
            return idx
    raise AssertionError(f"{ws.title}: missing header containing {required_terms}")


def column_values(ws, col, min_row=2):
    return [ws.cell(row, col).value for row in range(min_row, ws.max_row + 1) if ws.cell(row, col).value is not None]


def col_letter(col_idx):
    from openpyxl.utils import get_column_letter

    return get_column_letter(col_idx)


def validation_covers_column(ws, col_idx, required_tokens):
    col = col_letter(col_idx)
    for dv in ws.data_validations.dataValidation:
        formula = normalize_text(str(dv.formula1 or ""))
        if not all(normalize_text(token) in formula for token in required_tokens):
            continue
        for cell_range in dv.cells.ranges:
            if cell_range.min_col <= col_idx <= cell_range.max_col and cell_range.max_row >= 2:
                return True
        if col in str(dv.sqref):
            return True
    return False


def conditional_format_covers_column(ws, col_idx, required_tokens=()):
    for cf_range in ws.conditional_formatting:
        for rule in ws.conditional_formatting[cf_range]:
            formula = normalize_text(",".join(rule.formula or []))
            text = normalize_text(str(getattr(rule, "text", "") or ""))
            haystack = f"{formula} {text}"
            if required_tokens and not any(normalize_text(token) in haystack for token in required_tokens):
                continue
            sqref = str(getattr(cf_range, "sqref", cf_range))
            for ref in sqref.split():
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                if min_col <= col_idx <= max_col and max_row >= 2:
                    return True
    return False


def open_status_conditional_format_covers_column(ws, col_idx):
    for cf_range in ws.conditional_formatting:
        for rule in ws.conditional_formatting[cf_range]:
            formula = normalize_text(",".join(rule.formula or []))
            text = normalize_text(str(getattr(rule, "text", "") or ""))
            haystack = f"{formula} {text}"
            if "open" not in haystack and "closed" not in haystack:
                continue
            sqref = str(getattr(cf_range, "sqref", cf_range))
            for ref in sqref.split():
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                if min_col <= col_idx <= max_col and max_row >= 2:
                    return True
    return False


def assert_header_style(ws):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value is None:
            continue
        assert cell_fill_rgb(cell) == EXPECT["header_fill"]
        assert cell_font_rgb(cell) == EXPECT["header_font"]
        assert cell.font.bold


def assert_body_font(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.row == 1:
                continue
            assert cell.font.name == EXPECT["body_font"]
            assert int(cell.font.sz) == EXPECT["body_font_size"]


def iter_range_cells(wb, ref):
    sheet_name, coord = ref.split("!")
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            yield cell


def validation_refs(ws):
    out = set()
    for dv in ws.data_validations.dataValidation:
        out.add((str(dv.sqref), dv.formula1))
    return out


def conditional_format_specs(ws):
    specs = []
    for cf_range in ws.conditional_formatting:
        sqref = str(getattr(cf_range, "sqref", cf_range))
        for rule in ws.conditional_formatting[cf_range]:
            specs.append(
                {
                    "sqref": sqref,
                    "type": rule.type,
                    "operator": getattr(rule, "operator", None),
                    "formula": ",".join(rule.formula or []),
                    "text": getattr(rule, "text", None),
                }
            )
    return specs


def chart_reference(chart, attr):
    series = chart.series[0]
    ref = getattr(series, attr)
    if ref is None:
        return None
    if getattr(ref, "numRef", None) is not None:
        return ref.numRef.f
    if getattr(ref, "strRef", None) is not None:
        return ref.strRef.f
    return None


def test_output_exists_and_is_xlsx():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)
    assert OUTPUT_PATH != INPUT_PATH


def test_public_sheet_order_and_privacy_cleanup():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    assert all(ws.sheet_state == "visible" for ws in wb.worksheets)
    text = normalize_text(sheet_text(wb))
    hits = [phrase for phrase in EXPECT["forbidden_phrases"] if normalize_text(phrase) in text]
    assert not hits, f"Forbidden phrases still present: {hits}"
    assert normalize_text("compliance-review@city.example") in text


def test_data_sheets_tables_styles_and_protection():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert bool(wb.security.lockStructure)
    for sheet in ["Clean Intake", "Exception Register", "Renewal Calendar", "Audit Log"]:
        ws = wb[sheet]
        assert ws.protection.sheet, f"{sheet}: sheet is not protected"
        assert ws.max_row >= 6, f"{sheet}: expected repaired data rows"
        assert len(ws.tables) >= 1, f"{sheet}: missing Excel table object"
        assert_header_style(ws)
        assert_body_font(ws)


def test_risk_status_validation_comments_and_conditional_formatting():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    er = wb["Exception Register"]
    risk_col = find_header_col(er, ["risk"])
    status_col = find_header_col(er, ["status"])
    assert validation_covers_column(er, risk_col, ["High", "Medium", "Low"])
    assert validation_covers_column(er, status_col, ["Open"])
    assert conditional_format_covers_column(er, risk_col, ["High"])
    assert open_status_conditional_format_covers_column(er, status_col)

    renewal = wb["Renewal Calendar"]
    days_col = find_header_col(renewal, ["days"])
    try:
        bucket_col = find_header_col(renewal, ["bucket"])
    except AssertionError:
        bucket_col = find_header_col(renewal, [], ["status"])
    assert any(isinstance(value, str) and value.startswith("=") for value in column_values(renewal, days_col))
    assert any(isinstance(value, str) and value.startswith("=") for value in column_values(renewal, bucket_col))
    assert conditional_format_covers_column(renewal, days_col) or conditional_format_covers_column(renewal, bucket_col)

    comments = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment is not None:
                    comments.append(normalize_text(cell.comment.text))
    assert comments, "Workbook should include comments explaining key formula/control assumptions"
    assert any("formula" in comment or "count" in comment for comment in comments)
    assert any("owner" in comment or "contact" in comment or "status" in comment for comment in comments)


def test_dashboard_formulas_chart_and_workflow_semantics():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    ws = wb["Dashboard"]
    formulas = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    assert len(formulas) >= 4
    assert any("Exception Register" in formula or "ExceptionRegister" in formula or "Exception" in formula for formula in formulas)
    assert any("Renewal Calendar" in formula or "RenewalCalendar" in formula or "Renewal" in formula for formula in formulas)
    assert len(ws._charts) == 1
    chart = ws._charts[0]
    assert isinstance(chart, BarChart)
    dashboard_text = normalize_text(sheet_text(ws))
    assert all(label in dashboard_text for label in ["high", "medium", "low"])

    ci = wb["Clean Intake"]
    contact_col = find_header_col(ci, [], ["contact", "mailbox", "email"])
    contacts = [normalize_text(str(value)) for value in column_values(ci, contact_col)]
    assert contacts and all("compliance-review@city.example" in value for value in contacts)

    er = wb["Exception Register"]
    owner_col = find_header_col(er, ["owner"])
    owners = [normalize_text(str(value)) for value in column_values(er, owner_col)]
    assert owners
    assert all("@" not in owner and not any(char.isdigit() for char in owner) for owner in owners)

    audit_text = normalize_text(sheet_text(wb["Audit Log"]))
    assert "contact" in audit_text or "personal" in audit_text or "pii" in audit_text
    assert "owner" in audit_text or "individual" in audit_text
    assert "validation" in audit_text or "conditional" in audit_text or "protection" in audit_text


def test_source_artifact_was_not_modified():
    source = load_workbook(INPUT_PATH, data_only=False)
    text = normalize_text(sheet_text(source))
    missing = [phrase for phrase in EXPECT["source_must_contain"] if normalize_text(phrase) not in text]
    assert not missing, f"Source artifact no longer contains expected defects: {missing}"
