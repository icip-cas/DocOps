import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

sys.path.insert(0, str(Path(__file__).parent))
from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(__import__("os").environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text())
INPUT_PATH = Path(META["input_path"])
OUTPUT_PATH = Path(META["output_path"])

RELEASE_ORDER = [
    "Dashboard",
    "Forecast Model",
    "Exception Review",
    "Assumptions",
    "Demand Raw",
    "Archive",
]
REQUIRED_MODEL_HEADERS = [
    "Month",
    "Base Units",
    "Actual Units",
    "Scenario Growth",
    "Forecast Units",
    "Price",
    "Revenue",
    "Capacity",
    "Capacity Flag",
]


def _key(value):
    return normalize_text(norm_cell(value))


def _header_map(ws, row=1):
    out = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if value not in (None, ""):
            out[_key(value)] = col
    return out


def _find_header(headers, accepted, label):
    for item in accepted:
        norm = _key(item)
        if norm in headers:
            return headers[norm]
    raise AssertionError(f"Missing {label}; accepted headers: {accepted}")


def _rows_from_sheet(ws):
    headers = [_key(cell.value) for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, len(headers) + 1)]
        if any(value not in (None, "") for value in values):
            rows.append(dict(zip(headers, values)))
    return rows


def _formula(cell, label, required_terms=(), any_terms=()):
    value = str(cell.value or "").strip()
    assert value.startswith("="), f"{label}: expected a formula, found {cell.value!r}"
    norm = normalize_formula(value)
    missing = [term for term in required_terms if term.upper() not in norm]
    assert not missing, f"{label}: formula missing required terms {missing}; formula={value!r}"
    if any_terms:
        assert any(term.upper() in norm for term in any_terms), (
            f"{label}: formula should use one of {any_terms}; formula={value!r}"
        )
    return norm


def _compact_header(text):
    return re.sub(r"[^A-Z0-9]", "", str(text or "").upper())


def _has_cell_ref(formula_norm, col, row, header=None):
    if f"{get_column_letter(col).upper()}{row}" in formula_norm:
        return True
    if header:
        compact_formula = re.sub(r"[^A-Z0-9]", "", formula_norm)
        return _compact_header(header) in compact_formula
    return False


def _direct_demand_value(value, source_wb):
    text = str(value or "").strip()
    if not text.startswith("="):
        return value
    match = re.fullmatch(r"DEMANDRAW!([A-Z]+)(\d+)", normalize_formula(text))
    if not match:
        return value
    col, row = match.groups()
    return source_wb["Demand Raw"][f"{col}{row}"].value


def _cell_set_for_ranges(sqrefs):
    return set(expand_sqref_cells(sqrefs))


def _validation_covers_assumption_scenario(ws):
    covered = _cell_set_for_ranges(sheet_data_validation_ranges(ws))
    assert "B5" in covered, f"Assumptions!B5 should have scenario validation; found {sorted(covered)}"


def _conditional_formatting_covers(ws, row, min_col, max_col):
    for cf_range in all_cf_ranges(ws):
        text = str(cf_range)
        for token in re.split(r"\s+", text.replace("<ConditionalFormatting ", "").replace(">", "")):
            if not token or ":" not in token:
                continue
            try:
                min_c, min_r, max_c, max_r = range_boundaries(token)
            except ValueError:
                continue
            if min_r <= row <= max_r and min_c <= max_col and max_c >= min_col:
                return True
    return False


def _row_highlighted(ws, row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        rgb = (cell_fill_rgb(ws.cell(row, col)) or "").upper()
        if rgb and rgb not in {"FFFFFF", "000000", "00000000"}:
            return True
    return _conditional_formatting_covers(ws, row, min_col, max_col)


def _expected_breaches(source_wb):
    assumptions = source_wb["Assumptions"]
    scenario = norm_cell(assumptions["B5"].value)
    scenario_col = None
    for col in range(2, assumptions.max_column + 1):
        if norm_cell(assumptions.cell(1, col).value) == scenario:
            scenario_col = col
            break
    assert scenario_col is not None, f"Scenario {scenario!r} not found in Assumptions headers"
    growth = float(assumptions.cell(2, scenario_col).value)
    price = float(assumptions.cell(3, scenario_col).value)
    capacity = float(assumptions.cell(4, scenario_col).value)
    rows = []
    for row in _rows_from_sheet(source_wb["Demand Raw"]):
        base = float(row["base units"])
        forecast = base * (1 + growth)
        breach = forecast > capacity
        rows.append(
            {
                "month": norm_cell(row["month"]),
                "base": base,
                "actual": float(row["actual units"]),
                "growth": growth,
                "price": price,
                "capacity": capacity,
                "forecast": forecast,
                "breach": breach,
            }
        )
    return rows


def _dashboard_has_formula(ws, required_terms, any_terms, label):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if isinstance(value, str) and value.strip().startswith("="):
                norm = normalize_formula(value)
                if all(term.upper() in norm for term in required_terms) and any(
                    term.upper() in norm for term in any_terms
                ):
                    return True
    raise AssertionError(f"Dashboard missing formula for {label}")


def _exception_month_from_cell(cell, model_month_by_row):
    value = cell.value
    text = str(value or "").strip()
    if not text.startswith("="):
        return norm_cell(value)
    norm = normalize_formula(text)
    if "FORECASTMODEL[MONTH]" in norm and "CAPACITYFLAG" in norm and "BREACH" in norm:
        return "__DYNAMIC_BREACH_FORMULA__"
    flag_match = re.search(r"FORECASTMODEL!I(\d+).*BREACH", norm)
    month_match = re.search(r"FORECASTMODEL!A(\d+)", norm)
    if flag_match and month_match and flag_match.group(1) == month_match.group(1):
        return f"__DIRECT_BREACH_ROW_{month_match.group(1)}__"
    if "AGGREGATE" in norm and "BREACH" in norm and "FORECASTMODEL!A" in norm:
        return "__DYNAMIC_BREACH_FORMULA__"
    match = re.search(r"FORECASTMODEL!A(\d+)", norm)
    if match:
        return model_month_by_row.get(int(match.group(1)), "")
    return ""


def test_forecast_driver_cascade():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)

    source = load_workbook(INPUT_PATH, data_only=True)
    expected_rows = _expected_breaches(source)
    expected_by_month = {row["month"]: row for row in expected_rows}
    expected_breach_months = [row["month"] for row in expected_rows if row["breach"]]

    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == RELEASE_ORDER
    assert sheet_hidden_state_ok(wb["Demand Raw"], "hidden")
    assert sheet_hidden_state_ok(wb["Archive"], "hidden")
    _validation_covers_assumption_scenario(wb["Assumptions"])

    model = wb["Forecast Model"]
    dashboard = wb["Dashboard"]
    exceptions = wb["Exception Review"]

    table_range = table_ref(model, "ForecastModel")
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    assert min_row == 1, f"ForecastModel should include the header row; found {table_range}"
    assert max_row - min_row == len(expected_rows), (
        f"ForecastModel should contain exactly {len(expected_rows)} data rows; found {table_range}"
    )

    headers = _header_map(model)
    for header in REQUIRED_MODEL_HEADERS:
        assert _key(header) in headers, f"Forecast Model missing column {header!r}"
    table_cols = set(range(min_col, max_col + 1))
    for header in REQUIRED_MODEL_HEADERS:
        assert headers[_key(header)] in table_cols, (
            f"Column {header!r} should be inside ForecastModel table {table_range}"
        )

    month_col = headers["month"]
    base_col = headers["base units"]
    actual_col = headers["actual units"]
    growth_col = headers["scenario growth"]
    forecast_col = headers["forecast units"]
    price_col = headers["price"]
    revenue_col = headers["revenue"]
    capacity_col = headers["capacity"]
    flag_col = headers["capacity flag"]

    found_months = Counter()
    model_month_by_row = {}
    for row in range(min_row + 1, max_row + 1):
        month = norm_cell(_direct_demand_value(model.cell(row, month_col).value, source))
        assert month in expected_by_month, f"Unexpected month in Forecast Model: {month!r}"
        model_month_by_row[row] = month
        found_months[month] += 1
        expected = expected_by_month[month]

        assert float(_direct_demand_value(model.cell(row, base_col).value, source)) == expected["base"]
        assert float(_direct_demand_value(model.cell(row, actual_col).value, source)) == expected["actual"]

        growth_formula = _formula(
            model.cell(row, growth_col),
            f"Scenario Growth row {row}",
            required_terms=["ASSUMPTIONS"],
            any_terms=["IF", "LOOKUP", "INDEX", "MATCH", "CHOOSE"],
        )
        assert "B5" in growth_formula or "SCENARIO" in growth_formula, (
            f"Scenario Growth row {row}: formula should depend on the editable scenario driver"
        )
        forecast_formula = _formula(model.cell(row, forecast_col), f"Forecast Units row {row}")
        assert _has_cell_ref(forecast_formula, base_col, row, "Base Units"), (
            f"Forecast Units row {row}: formula should use Base Units"
        )
        assert _has_cell_ref(forecast_formula, growth_col, row, "Scenario Growth"), (
            f"Forecast Units row {row}: formula should use Scenario Growth"
        )
        assert not _has_cell_ref(forecast_formula, actual_col, row, "Actual Units"), (
            f"Forecast Units row {row}: formula should not be based on Actual Units"
        )
        price_formula = _formula(model.cell(row, price_col), f"Price row {row}", required_terms=["ASSUMPTIONS"])
        assert "B5" in price_formula or "SCENARIO" in price_formula
        revenue_formula = _formula(model.cell(row, revenue_col), f"Revenue row {row}")
        assert _has_cell_ref(revenue_formula, forecast_col, row, "Forecast Units"), f"Revenue row {row}: formula should use Forecast Units"
        assert _has_cell_ref(revenue_formula, price_col, row, "Price"), f"Revenue row {row}: formula should use Price"
        capacity_formula = _formula(model.cell(row, capacity_col), f"Capacity row {row}", required_terms=["ASSUMPTIONS"])
        assert "B5" in capacity_formula or "SCENARIO" in capacity_formula
        flag_formula = _formula(model.cell(row, flag_col), f"Capacity Flag row {row}")
        assert _has_cell_ref(flag_formula, forecast_col, row, "Forecast Units"), f"Capacity Flag row {row}: formula should use Forecast Units"
        assert _has_cell_ref(flag_formula, capacity_col, row, "Capacity"), f"Capacity Flag row {row}: formula should use Capacity"

        highlighted = _row_highlighted(model, row, min_col, max_col)
        if expected["breach"]:
            assert highlighted, f"Capacity breach month {month} should be highlighted"

    assert found_months == Counter(expected_by_month.keys()), (
        f"ForecastModel should preserve every raw demand month; expected {Counter(expected_by_month.keys())}, found {found_months}"
    )

    title = " ".join(norm_cell(dashboard.cell(row, col).value) for row in range(1, 4) for col in range(1, 4))
    assert "forecast" in normalize_text(title) and "dashboard" in normalize_text(title)
    assert "stress" in normalize_text(title) or "scenario" in normalize_text(title) or "assumptions" in normalize_text(title)
    _dashboard_has_formula(dashboard, ["FORECAST"], ["SUM", "SUBTOTAL"], "total forecast revenue")
    _dashboard_has_formula(dashboard, ["FORECAST"], ["COUNT", "SUM"], "capacity breach count")

    ex_headers = _header_map(exceptions)
    ex_month_col = _find_header(ex_headers, ["Month"], "Exception Review month column")
    ex_issue_col = _find_header(ex_headers, ["Issue", "Status", "Capacity Issue"], "Exception Review issue/status column")
    _find_header(ex_headers, ["Forecast Units", "Units"], "Exception Review forecast units column")
    _find_header(ex_headers, ["Capacity"], "Exception Review capacity column")
    _find_header(ex_headers, ["Overage", "Breach Amount"], "Exception Review overage column")

    exception_months = []
    issue_text = ""
    for row in range(2, exceptions.max_row + 1):
        if not any(exceptions.cell(row, col).value not in (None, "") for col in range(1, exceptions.max_column + 1)):
            continue
        month = _exception_month_from_cell(exceptions.cell(row, ex_month_col), model_month_by_row)
        if month == "__DYNAMIC_BREACH_FORMULA__":
            if len(exception_months) >= len(expected_breach_months):
                continue
            month = expected_breach_months[len(exception_months)]
        elif month.startswith("__DIRECT_BREACH_ROW_"):
            model_row = int(re.search(r"(\d+)", month).group(1))
            month = model_month_by_row.get(model_row, "")
            if not expected_by_month.get(month, {}).get("breach"):
                continue
        assert month, f"Exception Review row {row}: missing breach month"
        exception_months.append(month)
        issue_text += " " + norm_cell(exceptions.cell(row, ex_issue_col).value)

    assert Counter(exception_months) == Counter(expected_breach_months), (
        f"Exception Review should contain one row per capacity breach month; "
        f"expected {expected_breach_months}, found {exception_months}"
    )
    assert "capacity" in normalize_text(issue_text) or "breach" in normalize_text(issue_text)
