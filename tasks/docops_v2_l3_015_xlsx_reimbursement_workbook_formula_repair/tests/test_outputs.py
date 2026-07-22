import json
import os
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart
from openpyxl.utils import range_boundaries

from verifier_utils import cell_fill_rgb, cell_font_rgb, normalize_formula, normalize_text, run_preflight

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]
INPUT_PATH = Path(os.environ.get("INPUT_PATH", META["input_path"]))
OUTPUT_PATH = Path(os.environ.get("OUTPUT_PATH", META["output_path"]))


def sheet_text(wb):
    parts = list(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
                if cell.comment is not None:
                    parts.append(cell.comment.text)
    return "\n".join(parts)


def tab_color(ws):
    color = ws.sheet_properties.tabColor
    return None if color is None or color.rgb is None else color.rgb.upper()[-6:]


def values_match(actual, expected):
    if expected == "" and actual is None:
        return True
    if isinstance(expected, str) and isinstance(actual, (datetime, date)):
        return actual.strftime("%Y-%m-%d") == expected
    if isinstance(expected, str) and expected.startswith("="):
        return normalize_formula(actual) == normalize_formula(expected)
    return actual == expected


def assert_table_values(ws, expected_rows):
    assert ws.max_row == len(expected_rows)
    for row_idx, expected_row in enumerate(expected_rows, 1):
        actual = [ws.cell(row_idx, col).value for col in range(1, len(expected_row) + 1)]
        mismatches = [
            (col_idx, a, e)
            for col_idx, (a, e) in enumerate(zip(actual, expected_row), 1)
            if not values_match(a, e)
        ]
        assert not mismatches, f"{ws.title} row {row_idx} mismatches: {mismatches}"


def assert_dashboard_cells(ws):
    for cell, expected in EXPECT["dashboard_cells"].items():
        assert values_match(ws[cell].value, expected), f"Dashboard {cell}: wrong value/formula"


def assert_header_style(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(1, col)
        assert cell_fill_rgb(cell) == EXPECT["header_fill"]
        assert cell_font_rgb(cell) == EXPECT["header_font"]
        assert cell.font.bold


def assert_body_font(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.row == 1:
                continue
            assert cell.font.name == EXPECT["body_font"]
            assert int(cell.font.sz) == EXPECT["body_font_size"]


def iter_range_cells(wb, ref):
    sheet_name, coord = ref.split("!")
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            yield cell


def validation_refs(ws):
    out = set()
    for dv in ws.data_validations.dataValidation:
        formula = dv.formula1 or ""
        if formula.startswith("="):
            formula = formula[1:]
        out.add((str(dv.sqref), formula))
    return out


def conditional_format_specs(ws):
    specs = []
    for cf_range in ws.conditional_formatting:
        sqref = str(getattr(cf_range, "sqref", cf_range))
        for rule in ws.conditional_formatting[cf_range]:
            specs.append(
                {
                    "sqref": sqref,
                    "type": rule.type,
                    "operator": getattr(rule, "operator", None),
                    "formula": ",".join(rule.formula or []),
                    "text": getattr(rule, "text", None),
                }
            )
    return specs


def chart_reference(chart, attr):
    series = chart.series[0]
    ref = getattr(series, attr)
    if ref is None:
        return None
    if getattr(ref, "numRef", None) is not None:
        return ref.numRef.f
    if getattr(ref, "strRef", None) is not None:
        return ref.strRef.f
    return None


def cf_matches(spec, expected):
    if all(spec.get(key) == value for key, value in expected.items()):
        return True
    if spec.get("sqref") != expected.get("sqref"):
        return False
    formula = normalize_formula("=" + (spec.get("formula") or ""))
    if expected.get("type") == "containsText" and expected.get("text"):
        text = str(expected["text"]).upper()
        return spec.get("type") == "expression" and text in formula
    if expected.get("type") == "cellIs" and expected.get("operator") == "greaterThan":
        threshold = str(expected.get("formula", ""))
        return spec.get("type") == "expression" and ">" in formula and threshold in formula
    return False


def test_output_exists_and_is_xlsx():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)
    assert OUTPUT_PATH != INPUT_PATH


def test_sheet_order_tabs_tables_formulas_and_privacy_cleanup():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    for sheet, color in EXPECT["tab_colors"].items():
        assert tab_color(wb[sheet]) == color
    text = normalize_text(sheet_text(wb))
    hits = [phrase for phrase in EXPECT["forbidden_phrases"] if normalize_text(phrase) in text]
    assert not hits, f"Forbidden phrases still present: {hits}"
    assert_dashboard_cells(wb["Repair Dashboard"])
    for sheet, expected_rows in EXPECT["tables"].items():
        assert_table_values(wb[sheet], expected_rows)
        assert_header_style(wb[sheet], len(expected_rows[0]))
        assert_body_font(wb[sheet])


def test_tables_formats_validation_comments_and_conditional_formatting():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    for sheet, pane in EXPECT["freeze_panes"].items():
        assert wb[sheet].freeze_panes == pane
    for sheet, spec in EXPECT["excel_tables"].items():
        assert spec["name"] in wb[sheet].tables
        assert wb[sheet].tables[spec["name"]].ref == spec["ref"]
    for ref in EXPECT["currency_cells"]:
        for cell in iter_range_cells(wb, ref):
            assert cell.number_format == EXPECT["currency_format"]
    for item in EXPECT["data_validations"]:
        formula = item["formula1"][1:] if item["formula1"].startswith("=") else item["formula1"]
        assert (item["sqref"], formula) in validation_refs(wb[item["sheet"]])
    for ref, text in EXPECT["comments"].items():
        sheet, cell = ref.split("!")
        comment = wb[sheet][cell].comment
        assert comment is not None
        assert comment.text == text
    for cf in EXPECT["conditional_formats"]:
        specs = conditional_format_specs(wb[cf["sheet"]])
        expected = {key: value for key, value in cf.items() if key != "sheet"}
        assert any(cf_matches(spec, expected) for spec in specs), (
            f"Missing conditional format {expected}; found {specs}"
        )


def test_dashboard_chart_and_protection():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    chart_spec = EXPECT["chart"]
    ws = wb[chart_spec["sheet"]]
    assert len(ws._charts) == 1
    chart = ws._charts[0]
    assert isinstance(chart, BarChart)
    assert chart_reference(chart, "cat") == chart_spec["categories"]
    assert chart_reference(chart, "val") == chart_spec["values"]
    assert bool(wb.security.lockStructure)
    for sheet in EXPECT["protected_sheets"]:
        assert wb[sheet].protection.sheet


def test_source_artifact_was_not_modified():
    source = load_workbook(INPUT_PATH, data_only=False)
    text = normalize_text(sheet_text(source))
    missing = [phrase for phrase in EXPECT["source_must_contain"] if normalize_text(phrase) not in text]
    assert not missing, f"Source artifact no longer contains expected defects: {missing}"
