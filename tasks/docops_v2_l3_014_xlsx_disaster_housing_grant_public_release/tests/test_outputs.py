import json
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart
from openpyxl.utils import range_boundaries

from verifier_utils import cell_fill_rgb, cell_font_rgb, normalize_formula, normalize_text, run_preflight

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]
INPUT_PATH = Path(os.environ.get("INPUT_PATH", META["input_path"]))
OUTPUT_PATH = Path(os.environ.get("OUTPUT_PATH", META["output_path"]))


def sheet_text(wb):
    parts = list(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
    return "\n".join(parts)


def tab_color(ws):
    color = ws.sheet_properties.tabColor
    return None if color is None or color.rgb is None else color.rgb.upper()[-6:]


def row_values(ws, row_idx, max_col):
    return [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]


def values_match(actual, expected):
    if isinstance(expected, str) and expected.startswith("="):
        return normalize_formula(actual) == normalize_formula(expected)
    return actual == expected


def assert_table_values(ws, expected_rows):
    assert ws.max_row == len(expected_rows), f"{ws.title}: wrong row count"
    assert ws.max_column >= len(expected_rows[0]), f"{ws.title}: missing columns"
    for row_idx, expected_row in enumerate(expected_rows, 1):
        actual_row = row_values(ws, row_idx, len(expected_row))
        mismatches = [
            (col_idx, actual, expected)
            for col_idx, (actual, expected) in enumerate(zip(actual_row, expected_row), 1)
            if not values_match(actual, expected)
        ]
        assert not mismatches, f"{ws.title} row {row_idx} mismatches: {mismatches}"


def assert_dashboard_cells(ws):
    title = normalize_text(" ".join(str(ws.cell(row, 1).value or "") for row in range(1, 4)))
    for term in ["disaster", "housing", "grant", "public"]:
        assert term in title, f"Funding Dashboard title missing {term!r}"
    metric_text = normalize_text(
        "\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None)
    )
    for term in ["total applications", "total eligible amount", "senior review", "appeal review", "average priority"]:
        assert term in metric_text, f"Funding Dashboard missing metric {term!r}"
    formulas = [cell.value for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    required_bits = ["COUNTA", "SUM", "COUNTIF", "AVERAGE"]
    formula_text = normalize_formula("\n".join(formulas)).lower()
    for bit in required_bits:
        assert bit.lower() in formula_text, f"Funding Dashboard missing {bit} formula"
    assert "eligibilityregister" in formula_text, "Funding Dashboard formulas must reference Eligibility Register"


def assert_header_style(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(1, col)
        if cell.value in (None, ""):
            continue
        assert cell_fill_rgb(cell) == EXPECT["header_fill"], f"{ws.title} {cell.coordinate}: wrong header fill"
        assert cell_font_rgb(cell) == EXPECT["header_font"], f"{ws.title} {cell.coordinate}: wrong header font color"
        assert cell.font.bold, f"{ws.title} {cell.coordinate}: header is not bold"


def assert_body_font(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.row == 1:
                continue
            assert cell.font.name == EXPECT["body_font"], f"{ws.title} {cell.coordinate}: wrong body font"
            assert int(cell.font.sz) == EXPECT["body_font_size"], f"{ws.title} {cell.coordinate}: wrong body font size"


def iter_range_cells(wb, ref):
    sheet_name, coord = ref.split("!")
    ws = wb[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            yield cell


def conditional_format_specs(ws):
    specs = []
    for cf_range in ws.conditional_formatting:
        sqref = getattr(cf_range, "sqref", cf_range)
        for rule in ws.conditional_formatting[cf_range]:
            specs.append(
                {
                    "sqref": str(sqref),
                    "type": rule.type,
                    "operator": getattr(rule, "operator", None),
                    "formula": ",".join(rule.formula or []),
                }
            )
    return specs


def header_map(ws):
    return {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}


def row_by_application(ws, app_id):
    app_col = header_map(ws).get("Application ID")
    assert app_col, f"{ws.title}: missing Application ID column"
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, app_col).value == app_id:
            return row
    raise AssertionError(f"{ws.title}: missing application {app_id}")


def assert_eligibility_register(ws):
    headers = header_map(ws)
    required = [
        "Application ID",
        "County",
        "Damage Category",
        "Repair Estimate",
        "Insurance Proceeds",
        "Income Band",
        "Vulnerable Household",
        "Appeal Flag",
        "Inspection Date",
        "Eligible Amount",
        "Priority Score",
        "Review Status",
    ]
    missing = [header for header in required if header not in headers]
    assert not missing, f"Eligibility Register missing headers {missing}"
    app_ids = [ws.cell(row, headers["Application ID"]).value for row in range(2, ws.max_row + 1)]
    assert len(app_ids) == 12, f"Eligibility Register should contain 12 deduplicated applications, got {len(app_ids)}"
    assert len(set(app_ids)) == 12, "Eligibility Register still contains duplicate Application IDs"
    for app_id in [f"H-{idx:03d}" for idx in range(101, 113)]:
        assert app_id in app_ids, f"Eligibility Register missing {app_id}"

    checks = {
        "H-104": {"County": "North Fork"},
        "H-109": {"Damage Category": "Major"},
        "H-112": {"Income Band": "Low"},
        "H-107": {"Inspection Date": "2026-05-21"},
    }
    for app_id, expected in checks.items():
        row = row_by_application(ws, app_id)
        for header, value in expected.items():
            actual = ws.cell(row, headers[header]).value
            if header == "Inspection Date":
                actual = str(actual)[:10]
            assert actual == value, f"{app_id} {header}: expected {value!r}, got {actual!r}"

    for row in range(2, ws.max_row + 1):
        formula = str(ws.cell(row, headers["Eligible Amount"]).value)
        assert formula.startswith("="), f"Eligible Amount row {row} must be a formula"
        normalized = normalize_formula(formula).lower()
        assert f"d{row}-e{row}" in normalized, (
            f"Eligible Amount row {row} must use Repair Estimate minus Insurance Proceeds"
        )
        assert "max(" in normalized and "min(" in normalized, f"Eligible Amount row {row} missing MAX/MIN cap logic"
        formula = str(ws.cell(row, headers["Priority Score"]).value)
        normalized = normalize_formula(formula).lower()
        for term in ["Destroyed", "Major", "Moderate", "Low"]:
            assert term.lower() in normalized, f"Priority Score row {row} missing {term} logic"
        assert "10" in normalized, f"Priority Score row {row} missing Minor/default score logic"
        formula = str(ws.cell(row, headers["Review Status"]).value)
        normalized = normalize_formula(formula).lower()
        for term in ["Senior Review", "Appeal Review", "Ready"]:
            assert term.lower().replace(" ", "") in normalized, f"Review Status row {row} missing {term}"


def assert_county_rollup(ws):
    text = normalize_text("\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None))
    for county in ["East Bay", "North Fork", "South Ridge", "West Lake"]:
        assert normalize_text(county) in text, f"County Rollup missing {county}"
    formulas = [str(cell.value) for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    formula_text = normalize_formula("\n".join(formulas)).lower()
    assert "eligibilityregister" in formula_text, "County Rollup formulas must reference Eligibility Register"
    assert "countif" in formula_text and "sumif" in formula_text, "County Rollup missing count/sum formulas"


def assert_appeal_queue(ws):
    text = normalize_text("\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None))
    formula_text = normalize_formula(text).lower()
    expected_refs = {"H-103": "A4", "H-105": "A6", "H-109": "A10"}
    for app_id, ref in expected_refs.items():
        assert normalize_text(app_id) in text or ref.lower() in formula_text, f"Appeal Queue missing {app_id}"
    assert "H-101" not in text and "H-102" not in text, "Appeal Queue includes non-appeal applications"


def assert_publication_controls(ws):
    text = normalize_text("\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None))
    for term in ["12000", "18000", "28000", "42000"]:
        assert term in text, f"Publication Controls missing assistance cap {term}"
    for term in ["North Fork", "Major", "Low"]:
        assert normalize_text(term) in text, f"Publication Controls missing corrected fact {term}"


def chart_reference(chart, attr):
    series = chart.series[0]
    ref = getattr(series, attr)
    if ref is None:
        return None
    if getattr(ref, "numRef", None) is not None:
        return ref.numRef.f
    if getattr(ref, "strRef", None) is not None:
        return ref.strRef.f
    return None


def test_output_exists_and_is_xlsx():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)
    assert OUTPUT_PATH != INPUT_PATH


def test_sheets_order_tabs_and_private_material_removed():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    for sheet in EXPECT["sheet_order"]:
        assert tab_color(wb[sheet]) is not None, f"{sheet}: missing public tab color"
    text = normalize_text(sheet_text(wb))
    hits = [phrase for phrase in EXPECT["forbidden_phrases"] if normalize_text(phrase) in text]
    assert not hits, f"Forbidden phrases still present: {hits}"


def test_public_tables_and_dashboard_formulas():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert_dashboard_cells(wb["Funding Dashboard"])
    assert_eligibility_register(wb["Eligibility Register"])
    assert_county_rollup(wb["County Rollup"])
    assert_appeal_queue(wb["Appeal Queue"])
    assert_publication_controls(wb["Publication Controls"])
    for sheet in ["Eligibility Register", "County Rollup", "Appeal Queue", "Publication Controls"]:
        ws = wb[sheet]
        assert_header_style(ws, ws.max_column)
        assert_body_font(ws)


def test_excel_tables_formats_freeze_and_conditional_formatting():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    for sheet, pane in EXPECT["freeze_panes"].items():
        assert wb[sheet].freeze_panes == pane
    for ref in EXPECT["currency_cells"]:
        for cell in iter_range_cells(wb, ref):
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                assert cell.number_format == EXPECT["currency_format"], f"{ref}: wrong currency format at {cell.coordinate}"
    for cf in EXPECT["conditional_formats"]:
        specs = conditional_format_specs(wb[cf["sheet"]])
        expected = {key: value for key, value in cf.items() if key != "sheet"}
        semantically_present = expected in specs or any(
            spec["sqref"] == expected["sqref"]
            and (
                spec["formula"] == expected["formula"]
                or spec["formula"].replace("$", "").upper() in {expected["formula"].upper(), "J2>0", "K2>=70"}
            )
            for spec in specs
        )
        assert semantically_present, f"Missing conditional format {expected}; found {specs}"


def test_dashboard_chart_references_county_rollup():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    ws = wb[EXPECT["chart"]["sheet"]]
    assert len(ws._charts) == 1
    chart = ws._charts[0]
    assert isinstance(chart, BarChart)
    assert len(chart.series) == 1
    assert chart_reference(chart, "cat") == EXPECT["chart"]["categories"]
    assert chart_reference(chart, "val") == EXPECT["chart"]["values"]


def test_protection_applied():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert bool(wb.security.lockStructure)
    for sheet in EXPECT["protected_sheets"]:
        assert wb[sheet].protection.sheet, f"{sheet} is not protected"


def test_source_artifact_was_not_modified():
    source = load_workbook(INPUT_PATH, data_only=False)
    text = normalize_text(sheet_text(source))
    missing = [phrase for phrase in EXPECT["source_must_contain"] if normalize_text(phrase) not in text]
    assert not missing, f"Source artifact no longer contains expected seed defects: {missing}"
