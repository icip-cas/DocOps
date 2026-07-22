import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

sys.path.insert(0, str(Path(__file__).parent))
from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(__import__("os").environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text())
INPUT_PATH = Path(META["input_path"])
OUTPUT_PATH = Path(META["output_path"])

RELEASE_ORDER = [
    "Close Summary",
    "Clean Transactions",
    "Exceptions",
    "Raw Transactions",
    "Vendor Map",
    "Rules",
    "Archive",
]
REQUIRED_CLEAN_HEADERS = [
    "Date",
    "Vendor",
    "Amount",
    "Receipt",
    "Memo",
    "Category",
    "Owner",
    "Receipt Required",
    "Review Flag",
    "Signed Amount",
]


def _key(value):
    return normalize_text(norm_cell(value))


def _header_map(ws, row=1):
    out = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row, col).value
        if value not in (None, ""):
            out[_key(value)] = col
    return out


def _rows_from_sheet(ws):
    headers = [_key(cell.value) for cell in ws[1]]
    rows = []
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, len(headers) + 1)]
        if any(value not in (None, "") for value in values):
            rows.append(dict(zip(headers, values)))
    return rows


def _source_signature(row):
    return (
        norm_cell(row["date"]),
        norm_cell(row["vendor"]),
        float(row["amount"]),
        norm_cell(row["receipt"]),
        norm_cell(row["memo"]),
    )


def _formula(value):
    return str(value or "").strip()


def _require_formula(cell, label, required_terms=(), any_terms=()):
    formula = _formula(cell.value)
    assert formula.startswith("="), f"{label}: expected a formula, found {cell.value!r}"
    norm = normalize_formula(formula)
    missing = [term for term in required_terms if term.upper() not in norm]
    assert not missing, f"{label}: formula missing required references/terms: {missing}; formula={formula!r}"
    if any_terms:
        assert any(term.upper() in norm for term in any_terms), (
            f"{label}: formula must use one of {any_terms}; formula={formula!r}"
        )
    return norm


def _contains_cell_ref(formula_norm, col, row):
    return f"{get_column_letter(col).upper()}{row}" in formula_norm


def _uses_structured_ref(formula_norm, header):
    token = header.upper().replace(" ", "")
    compact = formula_norm.replace(" ", "")
    return f"[@{token}]" in compact or f"[{token}]" in compact


def _cell_set_for_ranges(sqrefs):
    return set(expand_sqref_cells(sqrefs))


def _receipt_validation_covers(clean, receipt_col, first_row, last_row):
    covered = _cell_set_for_ranges(sheet_data_validation_ranges(clean))
    expected = {f"{get_column_letter(receipt_col)}{row}" for row in range(first_row, last_row + 1)}
    assert expected.issubset(covered), (
        f"Receipt validation should cover {sorted(expected)}, found {sorted(covered)}"
    )


def _conditional_formatting_covers(ws, row, min_col, max_col):
    for cf_range in all_cf_ranges(ws):
        text = str(cf_range)
        for token in re.split(r"\s+", text.replace("<ConditionalFormatting ", "").replace(">", "")):
            if not token or ":" not in token:
                continue
            try:
                min_c, min_r, max_c, max_r = range_boundaries(token)
            except ValueError:
                continue
            if min_r <= row <= max_r and min_c <= max_col and max_c >= min_col:
                return True
    return False


def _row_highlighted(ws, row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        rgb = (cell_fill_rgb(ws.cell(row, col)) or "").upper()
        if rgb and rgb not in {"FFFFFF", "000000", "00000000"}:
            return True
    return _conditional_formatting_covers(ws, row, min_col, max_col)


def _summary_value_cell(summary, label):
    label_norm = normalize_text(label)
    for row in range(1, summary.max_row + 1):
        for col in range(1, min(summary.max_column, 4) + 1):
            value = summary.cell(row, col).value
            if label_norm == normalize_text(norm_cell(value)):
                return summary.cell(row, col + 1)
    raise AssertionError(f"Close Summary: missing label {label!r}")


def _require_any_header(ws, accepted_names, label):
    headers = _header_map(ws)
    for name in accepted_names:
        key = _key(name)
        if key in headers:
            return headers[key]
    for actual, col in headers.items():
        for name in accepted_names:
            wanted = _key(name)
            if wanted and (wanted in actual or actual in wanted):
                return col
    raise AssertionError(
        f"{ws.title}: missing {label} column; accepted {accepted_names!r}; "
        f"found {[ws.cell(1, c).value for c in range(1, ws.max_column + 1)]}"
    )


def _direct_clean_ref(formula):
    match = re.fullmatch(r"\s*=\s*'?Clean Transactions'?!\$?([A-Z]+)\$?(\d+)\s*", str(formula), re.I)
    if not match:
        return None
    return column_index_from_string(match.group(1)), int(match.group(2))


def _resolved_or_none(clean, cell):
    value = cell.value
    if not (isinstance(value, str) and value.startswith("=")):
        return value
    ref = _direct_clean_ref(value)
    if ref:
        col, row = ref
        return clean.cell(row, col).value
    return None


def _issue_text(clean, cell, fallback_clean_row=None, review_col=None):
    value = cell.value
    if not (isinstance(value, str) and value.startswith("=")):
        return norm_cell(value)
    ref = _direct_clean_ref(value)
    if ref:
        col, row = ref
        target = clean.cell(row, col).value
        if isinstance(target, str) and target.startswith("="):
            return normalize_formula(target)
        return norm_cell(target)
    if fallback_clean_row and review_col:
        target = clean.cell(fallback_clean_row, review_col).value
        if isinstance(target, str) and target.startswith("="):
            return normalize_formula(target)
        return norm_cell(target)
    return normalize_formula(value)


def _expected_exception_signatures(raw_rows, vendor_rows, rule_rows):
    vendor_map = {norm_cell(r["vendor"]): r for r in vendor_rows}
    rules = {norm_cell(r["category"]): r for r in rule_rows}
    duplicate_counts = Counter((norm_cell(r["date"]), norm_cell(r["vendor"]), float(r["amount"])) for r in raw_rows)
    expected = []
    for row in raw_rows:
        vendor = norm_cell(row["vendor"])
        amount = float(row["amount"])
        mapped = vendor_map.get(vendor)
        category = norm_cell(mapped["category"]) if mapped else "Unmapped"
        rule = rules.get(category)
        receipt_required = norm_cell(rule["requires receipt"]) if rule else "Yes"
        threshold = float(rule["review threshold"]) if rule and rule["review threshold"] is not None else 0.0
        reasons = []
        if not mapped:
            reasons.append("unmapped")
        if receipt_required.lower() == "yes" and norm_cell(row["receipt"]).lower() == "no":
            reasons.append("missing receipt")
        if duplicate_counts[(norm_cell(row["date"]), vendor, amount)] > 1:
            reasons.append("duplicate")
        if threshold > 0 and abs(amount) > threshold:
            reasons.append("threshold")
        if reasons:
            expected.append((_source_signature(row), reasons))
    return expected


def test_month_close_over10():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)

    source = load_workbook(INPUT_PATH, data_only=True)
    raw_rows = _rows_from_sheet(source["Raw Transactions"])
    vendor_rows = _rows_from_sheet(source["Vendor Map"])
    rule_rows = _rows_from_sheet(source["Rules"])
    expected_source = Counter(_source_signature(row) for row in raw_rows)
    expected_exceptions = _expected_exception_signatures(raw_rows, vendor_rows, rule_rows)
    expected_exception_source = Counter(sig for sig, _reasons in expected_exceptions)

    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == RELEASE_ORDER
    assert sheet_hidden_state_ok(wb["Raw Transactions"], "hidden")
    assert sheet_hidden_state_ok(wb["Archive"], "hidden")

    clean = wb["Clean Transactions"]
    ex = wb["Exceptions"]
    summary = wb["Close Summary"]

    table_range = table_ref(clean, "CleanTransactions")
    min_col, min_row, max_col, max_row = range_boundaries(table_range)
    assert min_row == 1, f"CleanTransactions should include the header row; found {table_range}"
    assert max_row - min_row == len(raw_rows), (
        f"CleanTransactions should contain exactly {len(raw_rows)} data rows; found {table_range}"
    )

    clean_headers = _header_map(clean)
    for header in REQUIRED_CLEAN_HEADERS:
        assert _key(header) in clean_headers, f"Clean Transactions missing column {header!r}"
    table_cols = set(range(min_col, max_col + 1))
    for header in REQUIRED_CLEAN_HEADERS:
        assert clean_headers[_key(header)] in table_cols, (
            f"Column {header!r} should be inside CleanTransactions table {table_range}"
        )
    assert freeze_pane_ref(clean) == "A2"

    date_col = clean_headers["date"]
    vendor_col = clean_headers["vendor"]
    amount_col = clean_headers["amount"]
    receipt_col = clean_headers["receipt"]
    memo_col = clean_headers["memo"]
    category_col = clean_headers["category"]
    owner_col = clean_headers["owner"]
    receipt_required_col = clean_headers["receipt required"]
    review_col = clean_headers["review flag"]
    signed_col = clean_headers["signed amount"]
    first_data_row = min_row + 1
    last_data_row = max_row

    clean_source_counter = Counter()
    clean_row_by_signature = {}
    for row in range(first_data_row, last_data_row + 1):
        sig = (
            norm_cell(clean.cell(row, date_col).value),
            norm_cell(clean.cell(row, vendor_col).value),
            float(clean.cell(row, amount_col).value),
            norm_cell(clean.cell(row, receipt_col).value),
            norm_cell(clean.cell(row, memo_col).value),
        )
        clean_source_counter[sig] += 1
        clean_row_by_signature.setdefault(sig, []).append(row)

        category_formula = _require_formula(
            clean.cell(row, category_col),
            f"Category row {row}",
            required_terms=["VENDORMAP"],
            any_terms=["LOOKUP", "INDEX", "MATCH", "FILTER"],
        )
        assert _contains_cell_ref(category_formula, vendor_col, row) or _uses_structured_ref(category_formula, "Vendor"), (
            f"Category row {row}: formula should use the row vendor"
        )
        owner_formula = _require_formula(
            clean.cell(row, owner_col),
            f"Owner row {row}",
            required_terms=["VENDORMAP"],
            any_terms=["LOOKUP", "INDEX", "MATCH", "FILTER"],
        )
        assert _contains_cell_ref(owner_formula, vendor_col, row) or _uses_structured_ref(owner_formula, "Vendor"), (
            f"Owner row {row}: formula should use the row vendor"
        )
        receipt_formula = _require_formula(
            clean.cell(row, receipt_required_col),
            f"Receipt Required row {row}",
            required_terms=["RULES"],
            any_terms=["LOOKUP", "INDEX", "MATCH", "FILTER"],
        )
        assert _contains_cell_ref(receipt_formula, category_col, row) or _uses_structured_ref(receipt_formula, "Category"), (
            f"Receipt Required row {row}: formula should use the row category"
        )
        review_formula = _require_formula(
            clean.cell(row, review_col),
            f"Review Flag row {row}",
            required_terms=["RULES"],
            any_terms=["OR", "IF", "COUNTIF", "COUNTIFS", "RULES"],
        )
        assert any(
            _contains_cell_ref(review_formula, col, row)
            for col in [category_col, receipt_col, receipt_required_col, vendor_col, amount_col]
        ) or any(
            _uses_structured_ref(review_formula, header)
            for header in ["Category", "Receipt", "Receipt Required", "Vendor", "Amount"]
        ), f"Review Flag row {row}: formula should depend on the current transaction"
        signed_formula = _require_formula(clean.cell(row, signed_col), f"Signed Amount row {row}")
        assert _contains_cell_ref(signed_formula, amount_col, row) or _uses_structured_ref(signed_formula, "Amount"), (
            f"Signed Amount row {row}: formula should reference the Amount cell"
        )

    assert clean_source_counter == expected_source, (
        f"Clean Transactions must preserve every raw transaction exactly once; "
        f"expected {expected_source}, found {clean_source_counter}"
    )
    _receipt_validation_covers(clean, receipt_col, first_data_row, last_data_row)

    for sig, _reasons in expected_exceptions:
        rows = clean_row_by_signature.get(sig, [])
        assert rows, f"Expected exception transaction missing from Clean Transactions: {sig}"
        assert any(_row_highlighted(clean, row, min_col, max_col) for row in rows), (
            f"Review transaction should be highlighted in Clean Transactions: {sig}"
        )

    summary_title = " ".join(norm_cell(summary.cell(row, col).value) for row in range(1, 3) for col in range(1, 4))
    assert "close summary" in normalize_text(summary_title)
    for label, required_terms in {
        "Revenue": ["CLEAN"],
        "Expenses": ["CLEAN"],
        "Net": [],
    }.items():
        value_cell = _summary_value_cell(summary, label)
        _require_formula(value_cell, f"Close Summary {label}", required_terms=required_terms)

    ex_headers = _header_map(ex)
    for header in ["Vendor", "Amount"]:
        assert _key(header) in ex_headers, f"Exceptions missing column {header!r}"
    ex_vendor_col = ex_headers["vendor"]
    ex_issue_col = _require_any_header(ex, ["Issue", "Issue Detail", "Review Flag", "Exception Type", "Detail"], "issue")
    ex_amount_col = ex_headers["amount"]
    expected_exception_keys = Counter((sig[1], sig[2]) for sig, _reasons in expected_exceptions)
    expected_reasons_by_key = {}
    for sig, reasons in expected_exceptions:
        expected_reasons_by_key.setdefault((sig[1], sig[2]), set()).update(reasons)
    expected_records_in_clean_order = []
    for sig, reasons in expected_exceptions:
        rows = clean_row_by_signature.get(sig, [])
        expected_records_in_clean_order.append((rows[0] if rows else 10**6, sig, reasons))
    expected_records_in_clean_order.sort(key=lambda item: item[0])

    exception_counter = Counter()
    issue_text_by_key = {}
    for row in range(2, ex.max_row + 1):
        if not any(ex.cell(row, col).value not in (None, "") for col in range(1, ex.max_column + 1)):
            continue
        fallback = expected_records_in_clean_order[row - 2] if row - 2 < len(expected_records_in_clean_order) else None
        vendor_value = _resolved_or_none(clean, ex.cell(row, ex_vendor_col))
        amount_value = _resolved_or_none(clean, ex.cell(row, ex_amount_col))
        if (vendor_value in (None, "")) and fallback:
            vendor_value = fallback[1][1]
        if (amount_value in (None, "")) and fallback:
            amount_value = fallback[1][2]
        vendor = norm_cell(vendor_value)
        amount = float(amount_value)
        key = (vendor, amount)
        assert key in expected_exception_keys, (
            f"Exceptions contains a transaction that is not expected to need review: vendor={vendor}, amount={amount}"
        )
        exception_counter[key] += 1
        issue_text_by_key.setdefault(key, "")
        fallback_clean_row = fallback[0] if fallback else None
        issue_text_by_key[key] += " " + _issue_text(clean, ex.cell(row, ex_issue_col), fallback_clean_row, review_col)

    assert exception_counter == expected_exception_keys, (
        f"Exceptions should contain one row per review transaction; "
        f"expected {expected_exception_keys}, found {exception_counter}"
    )
    for key, reasons in expected_reasons_by_key.items():
        issue_norm = normalize_text(issue_text_by_key.get(key, ""))
        for reason in reasons:
            assert reason in issue_norm or reason.split()[0] in issue_norm, (
                f"Exception row for {key} should mention {reason!r}; issue text={issue_text_by_key.get(key, '')!r}"
            )
