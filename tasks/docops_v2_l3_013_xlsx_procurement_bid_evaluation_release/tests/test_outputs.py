import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook

from verifier_utils import cell_fill_rgb, cell_font_rgb, normalize_formula, normalize_text, run_preflight

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]
INPUT_PATH = Path(os.environ.get("INPUT_PATH", META["input_path"]))
OUTPUT_PATH = Path(os.environ.get("OUTPUT_PATH", META["output_path"]))


def sheet_text(wb):
    parts = list(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
    return "\n".join(parts)


def workbook_values(wb):
    for ws in wb.worksheets:
        yield ws.title
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    yield str(cell.value)


def workbook_text(wb):
    return "\n".join(workbook_values(wb))


def row_values(ws, row_idx, max_col):
    return [ws.cell(row_idx, c).value for c in range(1, max_col + 1)]


def assert_table(ws, rows):
    assert ws.max_row == len(rows)
    cols = len(rows[0])
    for idx, row in enumerate(rows, 1):
        assert row_values(ws, idx, cols) == row, f"{ws.title} row {idx}: mismatch"


def tab_color(ws):
    c = ws.sheet_properties.tabColor
    return None if c is None or c.rgb is None else c.rgb.upper()[-6:]


def assert_summary(ws):
    text = normalize_text("\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None))
    assert "vendor" in text
    assert "responsive" in text
    assert "recommend" in text or "award" in text
    formulas = formula_cells(ws)
    assert formulas, "Evaluation Summary must contain formulas linked to rebuilt public sheets"


def assert_header_style(ws):
    header_cols = [c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None]
    assert header_cols, f"{ws.title}: missing header row"
    for col in header_cols:
        cell = ws.cell(1, col)
        assert cell_fill_rgb(cell) == EXPECT["header_fill"]
        assert cell_font_rgb(cell) == EXPECT["header_font"]
        assert cell.font.bold
    body_fonts = [
        cell.font.name
        for row in ws.iter_rows(min_row=2)
        for cell in row
        if cell.value is not None
    ]
    assert body_fonts, f"{ws.title}: missing public table body"
    assert any(font == EXPECT["body_font"] for font in body_fonts)


def validation_refs(ws):
    return {(str(dv.sqref), dv.formula1) for dv in ws.data_validations.dataValidation}


def formula_cells(ws):
    return [
        cell
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]


def sheet_norm(ws):
    return normalize_text(
        "\n".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None)
    )


def assert_sheet_terms(wb, sheet, terms):
    text = sheet_norm(wb[sheet])
    missing = [term for term in terms if normalize_text(term) not in text]
    assert not missing, f"{sheet}: missing required public terms {missing}"


def assert_any_sheet_terms(wb, terms):
    text = normalize_text(workbook_text(wb))
    missing = [term for term in terms if normalize_text(term) not in text]
    assert not missing, f"Workbook missing required corrected terms {missing}"


def assert_no_obsolete_corrections(wb):
    raw = workbook_text(wb)
    checks = {
        "VND-03A": r"(?<![A-Z0-9])VND-03A(?![A-Z0-9])",
        "Northstar Analytcs": r"\bNorthstar\s+Analytcs\b",
        "Leah Moor": r"\bLeah\s+Moor\b(?!e)",
        "Responsiv": r"\bResponsiv\b(?!e)",
    }
    hits = [label for label, pattern in checks.items() if re.search(pattern, raw, flags=re.IGNORECASE)]
    old_price_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, (int, float)) and value == 1240000:
                    old_price_cells.append(f"{ws.title}!{cell.coordinate}")
                elif isinstance(value, str) and re.fullmatch(r"\$?\s*1,?240,?000", value.strip()):
                    old_price_cells.append(f"{ws.title}!{cell.coordinate}")
    assert not hits, f"Obsolete correction values still present: {hits}"
    assert not old_price_cells, f"Old price $1,240,000 still present: {old_price_cells}"


def assert_public_tables_and_workflow(wb):
    assert_summary(wb["Evaluation Summary"])
    assert_sheet_terms(wb, "Vendor Register", ["Vendor ID", "Vendor Name", "VND-003", "Northstar Analytics", "Responsive"])
    assert_sheet_terms(wb, "Requirements Scoring", ["Vendor ID", "Requirement", "Weight", "Score"])
    assert_sheet_terms(wb, "Price Normalization", ["Vendor ID", "1204000"])
    assert_sheet_terms(wb, "Conflict-of-Interest Log", ["Vendor ID"])
    assert_sheet_terms(wb, "Award Recommendation", ["Vendor", "Recommendation"])
    assert_sheet_terms(wb, "Debrief Register", ["Vendor", "Debrief"])
    assert_sheet_terms(wb, "Data Validation Guide", ["VendorIDs", "B2:B16", "B2:B6"])
    assert_sheet_terms(wb, "Publication Style Guide", ["4F2D7F", "Aptos"])
    assert_sheet_terms(wb, "Appendix Index", ["Vendor", "Scoring"])
    assert_any_sheet_terms(wb, ["VND-003", "Northstar Analytics", "Leah Moore", "Responsive", "1204000"])


def assert_required_formulas(wb):
    for sheet in ["Evaluation Summary", "Requirements Scoring", "Price Normalization", "Award Recommendation"]:
        assert formula_cells(wb[sheet]), f"{sheet}: required formulas were flattened or omitted"


def header_named_columns(ws):
    columns = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is not None:
            columns[col] = normalize_text(value)
    return columns


def assert_number_formats(wb):
    for ws in wb.worksheets:
        headers = header_named_columns(ws)
        for col, header in headers.items():
            is_currency_col = (
                ("price" in header or "cost" in header or "amount" in header)
                and "score" not in header
                and "weight" not in header
                and "rank" not in header
            )
            if is_currency_col:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row, col)
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        assert cell.number_format == EXPECT["currency_format"], f"{ws.title}!{cell.coordinate}: wrong currency format"
            if re.search(r"\b(weight|percent)\b", header):
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row, col)
                    if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        assert cell.number_format == EXPECT["percent_format"], f"{ws.title}!{cell.coordinate}: wrong percent format"


def test_output_exists_and_is_xlsx():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)
    assert OUTPUT_PATH != INPUT_PATH


def test_sheets_tab_colors_and_private_material_removed():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    for sheet in EXPECT["sheet_order"]:
        assert tab_color(wb[sheet]) is not None, f"{sheet}: missing public tab color"
    text = normalize_text(sheet_text(wb))
    hits = [p for p in EXPECT["forbidden_phrases"] if normalize_text(p) in text]
    assert not hits, f"Forbidden phrases still present: {hits}"
    assert_no_obsolete_corrections(wb)


def test_public_tables_summary_and_formats():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert_public_tables_and_workflow(wb)
    assert_required_formulas(wb)
    for sheet in EXPECT["sheet_order"]:
        ws = wb[sheet]
        assert_header_style(ws)
        assert ws.freeze_panes == "A2", f"{sheet}: header row is not frozen"
        assert ws.auto_filter.ref, f"{sheet}: missing autofilter"
    assert_number_formats(wb)


def test_named_range_data_validation_and_hyperlinks():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    for name, expected in EXPECT["named_ranges"].items():
        defined_name = wb.defined_names[name]
        assert list(defined_name.destinations) == [("Vendor Register", "$A$2:$A$6")]
        assert expected.endswith("$A$2:$A$6")
    for item in EXPECT["data_validations"]:
        refs = validation_refs(wb[item["sheet"]])
        assert (item["sqref"], item["formula1"]) in refs
    for ref, target in EXPECT["hyperlinks"].items():
        sheet, cell = ref.split("!")
        assert wb[sheet][cell].hyperlink is not None
        assert str(wb[sheet][cell].hyperlink.target).lower().startswith("mailto:")


def test_source_artifact_was_not_modified():
    source = load_workbook(INPUT_PATH, data_only=False)
    assert "Evaluation_DRAFT_PRIVATE" in source.sheetnames
    assert "legal hold" in source.sheetnames
    text = normalize_text(sheet_text(source))
    assert "evaluator disagreement" in text
    assert "personal phone" in text
