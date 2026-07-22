import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook

from verifier_utils import cell_fill_rgb, cell_font_rgb, normalize_formula, normalize_text, run_preflight

META_PATH = Path(os.environ.get("TASK_METADATA_PATH", "/tests/task_metadata.json"))
if not META_PATH.exists():
    META_PATH = Path(__file__).parent / "task_metadata.json"
META = json.loads(META_PATH.read_text(encoding="utf-8"))
EXPECT = META["verifier_expectations"]
INPUT_PATH = Path(os.environ.get("INPUT_PATH", META["input_path"]))
OUTPUT_PATH = Path(os.environ.get("OUTPUT_PATH", META["output_path"]))


def sheet_text(wb):
    parts = list(wb.sheetnames)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    parts.append(str(cell.value))
    return "\n".join(parts)


def row_values(ws, row_idx, max_col):
    return [ws.cell(row_idx, col).value for col in range(1, max_col + 1)]


def norm_value(value):
    if isinstance(value, str):
        return " ".join(value.split())
    return value


def equivalent_value(actual, expected):
    actual = norm_value(actual)
    expected = norm_value(expected)
    if isinstance(expected, str) and expected.startswith("="):
        return normalize_formula(actual) == normalize_formula(expected)
    try:
        if actual is not None and expected is not None:
            return abs(float(actual) - float(expected)) < 1e-9
    except (TypeError, ValueError):
        pass
    return actual == expected


def correction_value_matches(field, actual, expected):
    if equivalent_value(actual, expected):
        return True
    if field == "Staging Time":
        actual_norm = normalize_text(str(actual))
        expected_norm = normalize_text(str(expected))
        if expected_norm.startswith("ambulance staging "):
            return actual_norm == expected_norm.replace("ambulance staging ", "")
    return False


def assert_table(ws, expected_rows, allow_style_header_variant=False):
    max_col = len(expected_rows[0])
    actual_header = row_values(ws, 1, max_col)
    if allow_style_header_variant:
        assert actual_header[0] == expected_rows[0][0]
        assert str(actual_header[1]).lower() in {"requirement", "public iap standard"}
    else:
        assert actual_header == expected_rows[0]
    assert ws.max_row == len(expected_rows), f"{ws.title}: wrong row count"
    for r_idx, expected in enumerate(expected_rows[1:], start=2):
        actual = row_values(ws, r_idx, max_col)
        if ws.title == "Correction Log":
            matches = [
                correction_value_matches(actual[1], a, e)
                for a, e in zip(actual, expected)
            ]
        else:
            matches = [equivalent_value(a, e) for a, e in zip(actual, expected)]
        assert all(matches), (
            f"{ws.title} row {r_idx}: expected {expected!r}, found {actual!r}"
        )

    for row_idx in range(1, ws.max_row + 1):
        extras = [ws.cell(row_idx, col).value for col in range(max_col + 1, ws.max_column + 1)]
        assert not any(value is not None for value in extras), f"{ws.title} row {row_idx}: unexpected extra values {extras!r}"


def formulas(ws):
    return [
        normalize_formula(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]


def assert_any_formula(formulas, required_parts, label):
    for formula in formulas:
        if all(part.upper() in formula for part in required_parts):
            return
    raise AssertionError(f"Summary: missing {label} formula linked to public workbook tables")


def assert_summary_content(ws):
    text = normalize_text(sheet_text(load_workbook(OUTPUT_PATH, data_only=False)))
    assert "city heat emergency" in text
    assert "2026-07-15 0600-1800" in text
    all_formulas = formulas(ws)
    assert_any_formula(all_formulas, ["BRANCHASSIGNMENTS!F"], "total assigned staff")
    assert_any_formula(all_formulas, ["MEDICALPLAN"], "medical plan rollup")
    assert_any_formula(all_formulas, ["COMMUNICATIONSPLAN"], "communications rollup")


def assert_header_style(ws, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(1, col)
        assert cell_fill_rgb(cell) == EXPECT["header_fill"], f"{ws.title}!{cell.coordinate}: wrong header fill"
        assert cell_font_rgb(cell) == EXPECT["header_font"], f"{ws.title}!{cell.coordinate}: wrong header font"
        assert cell.font.bold, f"{ws.title}!{cell.coordinate}: header not bold"
    assert ws["A2"].font.name == EXPECT["body_font"], f"{ws.title}: body font should be Calibri"


def test_output_exists_and_is_xlsx():
    run_preflight(META, INPUT_PATH, OUTPUT_PATH)
    assert OUTPUT_PATH != INPUT_PATH


def test_sheets_and_private_material_removed():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert wb.sheetnames == EXPECT["sheet_order"]
    text = normalize_text(sheet_text(wb))
    hits = [p for p in EXPECT["forbidden_phrases"] if normalize_text(p) in text]
    assert not hits, f"Forbidden phrases still present: {hits}"


def test_public_iap_tables_and_corrections():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert_table(wb["Incident Objectives"], EXPECT["objectives_rows"])
    assert_table(wb["Organization Assignments"], EXPECT["organization_rows"])
    assert_table(wb["Branch Assignments"], EXPECT["branch_rows"])
    assert_table(wb["Communications Plan"], EXPECT["communications_rows"])
    assert_table(wb["Medical Plan"], EXPECT["medical_rows"])
    assert_table(wb["Correction Log"], EXPECT["correction_rows"])
    assert_table(wb["Style Guide"], EXPECT["style_rows"], allow_style_header_variant=True)


def test_summary_formulas_and_structure():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert_summary_content(wb["Summary"])
    expected_filters = {
        "Incident Objectives": "A1:E6",
        "Organization Assignments": "A1:D7",
        "Branch Assignments": "A1:H7",
        "Communications Plan": "A1:F7",
        "Medical Plan": "A1:E5",
        "Correction Log": "A1:E6",
        "Style Guide": "A1:B6"
    }
    for sheet, ref in expected_filters.items():
        assert wb[sheet].freeze_panes == "A2", f"{sheet}: header row should be frozen"
        assert wb[sheet].auto_filter.ref == ref, f"{sheet}: wrong autofilter range"


def test_style_migration_applied():
    wb = load_workbook(OUTPUT_PATH, data_only=False)
    assert_header_style(wb["Incident Objectives"], 5)
    assert_header_style(wb["Organization Assignments"], 4)
    assert_header_style(wb["Branch Assignments"], 8)
    assert_header_style(wb["Communications Plan"], 6)
    assert_header_style(wb["Medical Plan"], 5)
    assert_header_style(wb["Correction Log"], 5)
    assert_header_style(wb["Style Guide"], 2)
    assert cell_fill_rgb(wb["Summary"]["A1"]) == EXPECT["header_fill"]
    assert wb["Summary"]["A1"].font.bold


def test_source_artifact_was_not_modified():
    source = load_workbook(INPUT_PATH, data_only=False)
    assert "IAP_DRAFT_2026-07-15" in source.sheetnames
    assert "mayor notes" in source.sheetnames
    text = normalize_text(sheet_text(source))
    assert "vendor liability" in text
    assert "personal phone" in text
