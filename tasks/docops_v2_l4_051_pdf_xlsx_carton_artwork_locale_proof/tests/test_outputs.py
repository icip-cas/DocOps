import json
import os
import re
import sys
from pathlib import Path
from datetime import date, datetime

import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from pypdf import PdfReader

sys.path.insert(0, str(Path(__file__).parent))
from verifier_utils import *  # noqa: F401,F403

META_PATH = Path(os.environ.get('TASK_METADATA_PATH', '/tests/task_metadata.json'))
META = json.loads(META_PATH.read_text(encoding='utf-8'))
EXPECT = META['verifier_expectations']


def _vopt_norm_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime('%Y-%m-%d %H:%M')
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value)
    text = text.replace('\u2013', '-').replace('\u2014', '-').replace('\u2212', '-')
    text = text.replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def _vopt_norm_formula(value):
    return re.sub(r'\s+', '', _vopt_norm_text(value)).upper()


def _vopt_assert_formula(actual, expected, label='formula'):
    assert _vopt_norm_formula(actual) == _vopt_norm_formula(expected), f"{label}: expected {expected!r}, found {actual!r}"


def _vopt_norm_rows(rows):
    return [[_vopt_norm_text(cell) for cell in row] for row in rows]


def _vopt_rows_equal(actual, expected):
    return _vopt_norm_rows(actual) == _vopt_norm_rows(expected)


def _vopt_rows_text(rows):
    return '\n'.join('|'.join(_vopt_norm_text(cell) for cell in row) for row in rows)


def _vopt_norm_key(value):
    return re.sub(r'[^a-z0-9]+', '', _vopt_norm_text(value).lower())


def _vopt_header_has(header, required):
    actual = {_vopt_norm_text(cell).lower() for cell in header}
    missing = [cell for cell in required if _vopt_norm_text(cell).lower() not in actual]
    assert not missing, f"Missing expected header columns {missing!r}; actual={header!r}"


def _vopt_section_text(section_part):
    return '\n'.join(p.text for p in getattr(section_part, 'paragraphs', []) if p.text is not None)


def _vopt_header_text(doc):
    return '\n'.join(_vopt_section_text(section.header) for section in doc.sections)


def _vopt_footer_text(doc):
    return '\n'.join(_vopt_section_text(section.footer) for section in doc.sections)


def _vopt_table_ref(ws, expected_name=None):
    tables = ws.tables
    if expected_name:
        for name in tables.keys():
            if str(name).lower() == str(expected_name).lower():
                return tables[name].ref
    vals = list(tables.values())
    assert vals, f"No Excel native table found on sheet {ws.title!r}"
    return vals[0].ref


def _vopt_clean_area(value):
    text = _vopt_norm_text(value).replace("'", '').replace('$', '').replace(' ', '')
    if '!' in text:
        text = text.split('!', 1)[1]
    return text.upper()


def _vopt_assert_print_area(ws, expected):
    actual = _vopt_clean_area(ws.print_area)
    target = _vopt_clean_area(expected)
    assert target in actual or actual in target, f"{ws.title}: expected print area {expected!r}, found {ws.print_area!r}"


def _vopt_freeze_pane(value):
    return getattr(value, 'coordinate', value)


def _vopt_validated_cells(ws):
    cells = set()
    for dv in ws.data_validations.dataValidation:
        for rng in dv.cells.ranges:
            text = str(rng).replace('$', '')
            try:
                min_col, min_row, max_col, max_row = range_boundaries(text)
            except ValueError:
                cells.add(text)
                continue
            if max_row > 10000:
                max_row = max(min_row, 200)
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cells.add(f"{get_column_letter(col)}{row}")
    return cells


class _VoptValidationRanges(list):
    def __init__(self, ws, ranges):
        super().__init__(ranges)
        self.ws = ws

    def __contains__(self, expected):
        expected = str(expected).replace('$', '')
        if super().__contains__(expected):
            return True
        try:
            min_col, min_row, max_col, max_row = range_boundaries(expected)
        except ValueError:
            return super().__contains__(expected)
        target = {f"{get_column_letter(col)}{row}" for row in range(min_row, max_row + 1) for col in range(min_col, max_col + 1)}
        return target.issubset(_vopt_validated_cells(self.ws))


def _vopt_dv_ranges(ws):
    ranges = []
    for dv in ws.data_validations.dataValidation:
        ranges.extend(str(rng).replace('$', '') for rng in dv.cells.ranges)
    return _VoptValidationRanges(ws, ranges)


def _vopt_ordered_subset(expected, actual):
    actual_iter = iter([_vopt_norm_text(item) for item in actual])
    for item in [_vopt_norm_text(value) for value in expected]:
        for candidate in actual_iter:
            if item == candidate:
                break
        else:
            return False
    return True



def _pdf_path():
    return Path(os.environ.get('PDF_OUTPUT_PATH', EXPECT['packet_output']))


def _xlsx_path():
    return Path(os.environ.get('XLSX_OUTPUT_PATH', EXPECT['workbook_output']))


def _pdf_texts(path):
    with pdfplumber.open(str(path)) as pdf:
        return [page.extract_text() or '' for page in pdf.pages]


def _norm_lower(value):
    return _vopt_norm_text(value).lower()


def _has_terms(text, terms):
    lowered = _norm_lower(text)
    return all(term.lower() in lowered for term in terms)


def _has_any(text, terms):
    lowered = _norm_lower(text)
    return any(term.lower() in lowered for term in terms)


def _has_bleed(text):
    normalized = _norm_lower(text)
    return '0.125' in normalized and 'bleed' in normalized


def _row_record_from_headers(ws, row):
    header = [_vopt_norm_text(ws.cell(2, col).value) for col in range(1, ws.max_column + 1)]
    return {_vopt_norm_key(header[col - 1]): ws.cell(row, col).value for col in range(1, ws.max_column + 1)}


def _assert_formula_contains(cell, tokens, context):
    formula = _vopt_norm_formula(cell.value)
    assert formula.startswith('='), f'{context}: expected live formula, found {cell.value!r}'
    missing = [token for token in tokens if token.upper() not in formula]
    assert not missing, f'{context}: formula missing {missing!r}; found {cell.value!r}'


def _qa_sheet(wb):
    candidates = []
    for ws in wb.worksheets:
        if ws.sheet_state != 'visible':
            continue
        text = _worksheet_text(ws)
        if 'sku-a12' in text and 'sku-b07' in text:
            candidates.append(ws)
    assert candidates, f'No visible QA worksheet with artwork SKUs found; sheets={wb.sheetnames!r}'
    return candidates[0]


def _worksheet_text(ws):
    values = []
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if value is not None:
                values.append(_vopt_norm_text(value))
    return '\n'.join(values).lower()


def _find_header_row(ws):
    for row in range(1, min(ws.max_row, 12) + 1):
        keys = [_vopt_norm_key(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if (
            'sku' in keys
            and any(key in keys for key in ['panel', 'panelname', 'requiredpage'])
            and any(key in keys for key in ['publishstatus', 'qastatus', 'result', 'includedinpdf', 'finalpdfincluded'])
        ):
            return row
    raise AssertionError('Could not locate QA checklist header row')


def _records_from_header(ws, header_row):
    headers = [_vopt_norm_key(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    records = []
    for row in range(header_row + 1, ws.max_row + 1):
        cells = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        row_text = ' '.join(_vopt_norm_text(value) for value in cells if value is not None)
        if not row_text:
            continue
        records.append({headers[col - 1]: cells[col - 1] for col in range(1, ws.max_column + 1) if headers[col - 1]})
    return records


def _record_text(record):
    return ' '.join(_vopt_norm_text(value) for value in record.values() if value is not None).lower()


def _assert_workbook_sku_records(ws):
    records = _records_from_header(ws, _find_header_row(ws))
    all_rows = '\n'.join(_record_text(record) for record in records)
    required = [
        ('SKU-A12 English Panel', ['sku-a12', 'english panel']),
        ('SKU-A12 Spanish Panel', ['sku-a12', 'spanish panel', 'warning']),
        ('SKU-B07 English Panel', ['sku-b07', 'english panel']),
        ('SKU-X99 Legacy French Panel', ['sku-x99', 'legacy french panel']),
    ]
    for label, terms in required:
        assert all(term in all_rows for term in terms), f'Missing QA checklist record for {label}'
    retired_rows = [text for text in (_record_text(record) for record in records) if 'sku-x99' in text]
    assert retired_rows, 'SKU-X99 retired row missing from QA checklist'
    assert any(
        ('do not publish' in text or 'excluded' in text or 'no' in text or 'internal' in text)
        and ('retired' in text or 'legacy french' in text)
        for text in retired_rows
    ), 'SKU-X99 row does not clearly mark retired/do-not-publish disposition'


def _assert_any_formulas(ws):
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append(cell.value)
    assert formulas, 'Artwork QA workbook must include live formulas'


def _assert_pdf_retired_mentions_are_exclusions(texts):
    for idx, text in enumerate(texts, start=1):
        lower = _norm_lower(text)
        if 'sku-x99' in lower or 'legacy french panel' in lower:
            assert any(term in lower for term in ['excluded', 'removed', 'retired', 'omitted', 'filtered', 'do not publish']), (
                f'PDF page {idx} mentions retired artwork without a clear exclusion context'
            )


def test_outputs_exist():
    assert _pdf_path().exists()
    assert _xlsx_path().exists()
    assert _pdf_path().suffix.lower() == '.pdf'
    assert _xlsx_path().suffix.lower() == '.xlsx'


def test_pdf_artwork_proof_pages_bookmarks_and_boundary():
    path = _pdf_path()
    reader = PdfReader(str(path))
    assert len(reader.pages) == EXPECT['pdf_page_count']
    texts = _pdf_texts(path)
    expected_pages = [
        ['carton', 'artwork', 'proof'],
        ['SKU-A12', 'English'],
        ['SKU-A12', 'Spanish'],
        ['SKU-B07', 'English'],
        ['QA', 'Certificate'],
    ]
    for terms, page_text in zip(expected_pages, texts):
        assert _has_terms(page_text, terms), f"Missing expected PDF page terms {terms!r}"
    full = '\n'.join(texts)
    _assert_pdf_retired_mentions_are_exclusions(texts)
    assert _has_terms(full, ['SKU-A12', 'Spanish']) and _has_any(full, ['warning', 'review']), 'Missing Spanish panel warning/review status'
    assert _has_any(full, ['excluded', 'removed', 'retired', 'filtered', 'omitted']), 'Missing retired artwork exclusion note'
    assert _has_bleed(full), 'Missing retained 0.125 inch bleed note'
    assert _has_terms(full, ['EN']) or _has_terms(full, ['English']), 'Missing EN locale'
    assert _has_terms(full, ['ES']) or _has_terms(full, ['Spanish']), 'Missing ES locale'
    outline = [title for _level, title in flatten_outline(reader.outline)]
    assert any(_has_terms(title, ['cover']) or _has_terms(title, ['carton', 'proof']) for title in outline), 'Missing cover bookmark'
    for terms in [['SKU-A12', 'English'], ['SKU-A12', 'Spanish'], ['SKU-B07', 'English'], ['QA', 'Certificate']]:
        assert any(_has_terms(title, terms) for title in outline), f'Missing bookmark for {terms!r}'


def test_xlsx_qa_controls():
    wb = load_workbook(_xlsx_path(), data_only=False)
    assert 'Rules' in wb.sheetnames, f'Missing hidden Rules sheet; sheets={wb.sheetnames!r}'
    ws = _qa_sheet(wb)
    _assert_workbook_sku_records(ws)
    table_ref = _vopt_table_ref(ws)
    assert table_ref, 'Artwork QA sheet must include a native Excel table'
    assert ws.print_area, 'Artwork QA sheet must define a print area'
    assert ws.freeze_panes, 'Artwork QA sheet must freeze panes'
    _assert_any_formulas(ws)
    actual_ranges = _vopt_dv_ranges(ws)
    assert actual_ranges, 'Artwork QA sheet must include data validation'
    assert len(ws.conditional_formatting) >= 1
    for sheet in EXPECT['hidden_sheets']:
        assert wb[sheet].sheet_state in ('hidden', 'veryHidden')
    names = {dn.name for dn in wb.defined_names.values()}
    assert names, 'Workbook must include defined names'


def test_cross_output_sku_consistency():
    pdf_text = '\n'.join(_pdf_texts(_pdf_path()))
    wb = load_workbook(_xlsx_path(), data_only=False)
    ws = _qa_sheet(wb)
    sheet_text = _worksheet_text(ws)
    for sku in EXPECT['public_skus']:
        assert sku.lower() in sheet_text, f'{sku} missing from QA checklist'
        assert sku in pdf_text, f'{sku} missing from PDF proof'
    for sku in EXPECT['retired_skus']:
        assert sku.lower() in sheet_text, f'{sku} missing from QA checklist retired/disposition record'
    _assert_pdf_retired_mentions_are_exclusions(_pdf_texts(_pdf_path()))
